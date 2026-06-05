"""Verify the rebuilt workbook matches the GPT-prompt spec sheet by sheet."""
from openpyxl import load_workbook
import pandas as pd

P = r"C:\Users\rohit.kaushik\Downloads\Unified Audit Tool\REBUILD_ADP_vs_UZIO_Withholding_JMParcel.xlsx"

SPEC_REQUIRED = [
    "Summary", "Field Mapping Rules", "Mismatch Summary",
    "Mismatches (All)", "Mismatches (Active)", "Mismatches (Terminated)",
    "Missing in UZIO", "Missing in ADP",
]
SPEC_MISMATCH_COLS = [
    "Employee ID", "Employee Name", "Employment Status",
    "Category",
    "Field Label", "Field Key",
    "ADP Raw Value", "UZIO Raw Value",
    "ADP Normalized Value", "UZIO Normalized Value",
    "Rule Applied",
    "ADP Column Name", "UZIO Column Name",
]
SPEC_SUMMARY_COLS = ["Category", "Field Label", "Field Key", "Mismatch Count", "Unique Employee Count"]
SPEC_RULES_COLS = ["Field Label", "Field Key", "ADP Column Found",
                   "UZIO Column Found", "Type", "Notes"]


def section(title): print("\n" + "=" * 72); print(title); print("=" * 72)


section("Sheet inventory vs spec")
wb = load_workbook(P, read_only=True, data_only=True)
present = wb.sheetnames
for n in SPEC_REQUIRED:
    print(f"  [{'OK' if n in present else 'MISSING':>7}]  {n}")
extra = [n for n in present if n not in SPEC_REQUIRED]
print(f"\n  Extra sheets (additive over spec): {extra}")
wb.close()


def check_cols(sheet, required):
    df = pd.read_excel(P, sheet_name=sheet, dtype=str)
    cols = list(df.columns)
    missing = [c for c in required if c not in cols]
    return cols, missing


section("Mismatches (All) columns vs spec")
cols, missing = check_cols("Mismatches (All)", SPEC_MISMATCH_COLS)
for c in SPEC_MISMATCH_COLS:
    print(f"  [{'OK' if c in cols else 'MISS':>4}]  {c}")
print(f"  Missing: {missing}")


section("Mismatch Summary columns vs spec")
cols, missing = check_cols("Mismatch Summary", SPEC_SUMMARY_COLS)
for c in SPEC_SUMMARY_COLS:
    print(f"  [{'OK' if c in cols else 'MISS':>4}]  {c}")
print(f"  Missing: {missing}")


section("Field Mapping Rules columns vs spec")
cols, missing = check_cols("Field Mapping Rules", SPEC_RULES_COLS)
for c in SPEC_RULES_COLS:
    print(f"  [{'OK' if c in cols else 'MISS':>4}]  {c}")
print(f"  Missing: {missing}")


section("Mismatches (All) data — check FIELD_LABEL is from YAML, not titlecase")
df = pd.read_excel(P, sheet_name="Mismatches (All)", dtype=str)
print(df[["Employee ID", "State Code", "Field Label", "Field Key",
          "ADP Column Name", "UZIO Column Name", "ADP Raw Value", "UZIO Raw Value"]].to_string(index=False))


section("Field Mapping Rules sample")
df = pd.read_excel(P, sheet_name="Field Mapping Rules", dtype=str)
print(df.to_string(index=False))


section("Sort check — Mismatches (All)")
df = pd.read_excel(P, sheet_name="Mismatches (All)", dtype=str)
expected = sorted(zip(df["Employee ID"], df["Field Label"]))
actual = list(zip(df["Employee ID"], df["Field Label"]))
print(f"  In Employee ID -> Field Label order: {actual == expected}")
