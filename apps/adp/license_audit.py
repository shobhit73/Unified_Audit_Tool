import streamlit as st
import pandas as pd
import io
import re
from datetime import datetime, date
import numpy as np

APP_TITLE = "ADP License Details Audit"

# --- HELPER FUNCTIONS ---
def norm_blank(x):
    """Normalize NaN, None, or completely whitespace strings to empty string."""
    if pd.isna(x) or x is None:
        return ""
    if isinstance(x, str):
        s = x.strip()
        if not s:
             return ""
        return s
    return x

def try_parse_date(x):
    """Safely parse a date string or object into YYYY-MM-DD string."""
    x = norm_blank(x)
    if x == "":
        return ""
    if isinstance(x, (datetime, date, np.datetime64, pd.Timestamp)):
        return pd.to_datetime(x).strftime('%Y-%m-%d')
    if isinstance(x, str):
        # Handle '1900-01-01 00:00:00' common Excel raw string formats
        s = x.strip().split(' ')[0]
        try:
            return pd.to_datetime(s, errors="raise").strftime('%Y-%m-%d')
        except Exception:
            return s
    return str(x)

def safe_read_excel(file, dtype=None, header=None):
    """
    Robust reader for Excel files. 
    Handles openpyxl 'Invalid datetime value' errors by falling back 
    to a direct openpyxl read to bypass automated parsing failures.
    """
    try:
        file.seek(0)
        return pd.read_excel(file, dtype=dtype, header=header)
    except Exception as e:
        err_msg = str(e)
        # If openpyxl fails due to date parsing or other data-specific issues
        if "Invalid datetime value" in err_msg or "datetime" in err_msg.lower():
            try:
                import openpyxl
                file.seek(0)
                # data_only=True retrieves values only, bypassing complex formula/date parsing
                wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
                try:
                    sheet = wb.active
                    if sheet is None:
                        sheet = wb[wb.sheetnames[0]]
                except:
                    sheet = wb[wb.sheetnames[0]]
                
                data = []
                for row in sheet.iter_rows(values_only=True):
                    # Manually convert each cell to string to match 'dtype=str' intent
                    data.append([str(c).strip() if c is not None else "" for c in row])
                
                df = pd.DataFrame(data)
                return df
            except Exception:
                # If manual read also fails, raise original error
                raise e
        raise e

def read_uzio_license(file) -> pd.DataFrame:
    """Reads UZIO license report, extracting exact headers while bypassing corrupt metadata."""
    try:
        df = safe_read_excel(file, header=None, dtype=str)
        # Find the header row by looking for 'Employee ID'
        header_idx = -1
        for i, row in df.head(20).iterrows():
            if any(str(c).strip() == 'Employee ID' for c in row.values if pd.notna(c)):
                header_idx = i
                break
        
        if header_idx != -1:
            df.columns = [str(c).strip() if pd.notna(c) else f"Unnamed_{i}" for i, c in enumerate(df.iloc[header_idx])]
            df = df.iloc[header_idx + 1:].reset_index(drop=True)
            df = df.replace('None', '')
            df.dropna(how='all', inplace=True)
            return df
        else:
            # Fallback if not found in first 20 rows
            file.seek(0)
            df = safe_read_excel(file, dtype=str)
            return df
    except Exception as e:
        st.error(f"Could not read Uzio file: {e}")
        return None

def read_adp_license(file) -> pd.DataFrame:
    """Reads ADP license report, extracting headers while bypassing metadata."""
    try:
        df = safe_read_excel(file, header=None, dtype=str)
        # Locate header row
        header_idx = -1
        for i, row in df.head(20).iterrows():
            if any(str(c).strip() == 'Associate ID' for c in row.values if pd.notna(c)):
                header_idx = i
                break
                
        if header_idx != -1:
            df.columns = [str(c).strip() if pd.notna(c) else f"Unnamed_{i}" for i, c in enumerate(df.iloc[header_idx])]
            df = df.iloc[header_idx + 1:].reset_index(drop=True)
            df = df.replace('None', '')
            df.dropna(how='all', inplace=True)
            return df
        else:
            # Fallback
            file.seek(0)
            df = safe_read_excel(file, dtype=str)
            return df
    except Exception as e:
        st.error(f"Could not read ADP file: {e}")
        return None

# --- AUDIT LOGIC ---
def run_license_audit(uzio_df, adp_df):
    """
    Compares Uzio and ADP License data based on:
      Uzio: Employee ID, License Number, License Expiration Date
      ADP: Associate ID, License/Certification ID, Expiration Date
    """
    UZIO_KEY = 'Employee ID'
    UZIO_NUM_COL = 'License Number'
    UZIO_DATE_COL = 'License Expiration Date'
    
    ADP_KEY = 'Associate ID'
    ADP_NUM_COL = 'License/Certification Code' if 'License/Certification Code' in adp_df.columns else 'License/Certification ID'
    ADP_DATE_COL = 'Expiration Date'
    
    required_uzio = [UZIO_KEY, UZIO_NUM_COL, UZIO_DATE_COL]
    required_adp = [ADP_KEY, ADP_NUM_COL, ADP_DATE_COL]
    
    missing_uzio = [c for c in required_uzio if c not in uzio_df.columns]
    if missing_uzio:
        st.error(f"Missing required Uzio columns: {', '.join(missing_uzio)}")
        return None
            
    missing_adp = [c for c in required_adp if c not in adp_df.columns]
    if missing_adp:
        st.error(f"Missing required ADP columns: {', '.join(missing_adp)}")
        return None
            
    # Normalize Keys
    uzio_df[UZIO_KEY] = uzio_df[UZIO_KEY].apply(lambda x: str(x).strip().lstrip('0') if norm_blank(x) != "" else "")
    adp_df[ADP_KEY] = adp_df[ADP_KEY].apply(lambda x: str(x).strip().lstrip('0') if norm_blank(x) != "" else "")
    
    all_keys = sorted(list(set(uzio_df[UZIO_KEY].unique()).union(set(adp_df[ADP_KEY].unique()))))
    all_keys = [k for k in all_keys if k]
    
    # Pre-process DataFrames into dictionaries of records for faster lookup
    uzio_records = uzio_df.to_dict('records')
    adp_records = adp_df.to_dict('records')
    
    uzio_map = {}
    for r in uzio_records:
        k = r[UZIO_KEY]
        if k:
            if k not in uzio_map:
                uzio_map[k] = []
            uzio_map[k].append(r)
            
    adp_map = {}
    for r in adp_records:
        k = r[ADP_KEY]
        if k:
            if k not in adp_map:
                adp_map[k] = []
            adp_map[k].append(r)
            
    rows = []
    
    for eid in all_keys:
        uzio_licenses = uzio_map.get(eid, [])
        adp_licenses = adp_map.get(eid, [])
        
        # We only care if Uzio has a license according to the logic requested
        for uz_rec in uzio_licenses:
            uz_num = str(uz_rec.get(UZIO_NUM_COL, "")).strip()
            uz_raw_date = uz_rec.get(UZIO_DATE_COL, "")
            uz_date = try_parse_date(uz_raw_date)
            
            if not uz_num:
                # Uzio license number is blank
                rows.append({
                    "Employee ID": eid,
                    "Field": "License Number",
                    "Status": "Missing in Uzio",
                    "Uzio Value": "",
                    "ADP Value": ""
                })
                # Check expiration date also if needed? Usually we flag the whole record.
                # If they want, we can flag date too, but blank license is the main issue.
                continue
                
            # Uzio has a license number. Check if it's in ADP.
            match_found = False
            matched_adp_rec = None
            
            for adp_rec in adp_licenses:
                adp_num = str(adp_rec.get(ADP_NUM_COL, "")).strip()
                if uz_num.lower() == adp_num.lower():
                    match_found = True
                    matched_adp_rec = adp_rec
                    break
                    
            if match_found:
                # License Number matched
                rows.append({
                    "Employee ID": eid,
                    "Field": "License Number",
                    "Status": "Data Match",
                    "Uzio Value": uz_num,
                    "ADP Value": str(matched_adp_rec.get(ADP_NUM_COL, "")).strip()
                })
                
                # Check Expiration Date
                adp_raw_date = matched_adp_rec.get(ADP_DATE_COL, "")
                adp_date = try_parse_date(adp_raw_date)
                
                status = "Data Match" if uz_date == adp_date else "Data Mismatch"
                rows.append({
                    "Employee ID": eid,
                    "Field": "Expiration Date",
                    "Status": status,
                    "Uzio Value": uz_date,
                    "ADP Value": adp_date
                })
            else:
                # Not found in ADP
                rows.append({
                    "Employee ID": eid,
                    "Field": "License Number",
                    "Status": "Missing in ADP",
                    "Uzio Value": uz_num,
                    "ADP Value": ""
                })

    return pd.DataFrame(rows)

# --- UI RENDER FLOW ---
def render_ui():
    st.title("ADP License Details Audit Tool")
    st.write("Compare license numbers and expiration dates between ADP and Uzio. Output will be generated as a single sheet.")

    client_name = st.text_input("Client Name", key="adp_license_client_name")

    st.markdown("### Step 1: Upload Files")
    col1, col2 = st.columns(2)
    with col1:
        uzio_file = st.file_uploader("Upload UZIO License Report (.xlsx)", type=['xlsx', 'csv'], key='uzio_license_upload')
    with col2:
        adp_file = st.file_uploader("Upload ADP License Report (.xlsx)", type=['xlsx', 'csv'], key='adp_license_upload')

    if uzio_file and adp_file:
        try:
            with st.spinner("Extracting Licenses..."):
                uzio_df = read_uzio_license(uzio_file)
                adp_df = read_adp_license(adp_file)

            if uzio_df is None or adp_df is None:
                st.error("Failed to parse one or both files.")
                return

            submit_audit = st.button("Run License Audit", type="primary")

            if submit_audit:
                 if not client_name:
                     st.warning("Please enter a Client Name before running the audit.")
                     return

                 with st.spinner("Running Audit & Generating Match Report..."):
                     result_df = run_license_audit(uzio_df, adp_df)
                     
                 if result_df is not None:
                      st.success("Audit Complete!")
                      
                      st.markdown("### Audit Summary")
                      
                      col1, col2, col3, col4 = st.columns(4)
                      counts = result_df['Status'].value_counts().to_dict()
                      
                      col1.metric("Total Match", counts.get("Data Match", 0))
                      col2.metric("Total Mismatch", counts.get("Data Mismatch", 0))
                      col3.metric("Missing in UZIO", counts.get("Missing in Uzio", 0))
                      col4.metric("Missing in ADP", counts.get("Missing in ADP", 0))

                      st.dataframe(result_df)

                      # Download Button
                      output_filename = f"{client_name.strip()}_License_Audit_Report_{datetime.now().strftime('%d_%m_%Y')}.xlsx"
                      
                      buffer = io.BytesIO()
                      with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                          result_df.to_excel(writer, sheet_name='License Audit Results', index=False)

                      st.download_button(
                          label="Download Excel Report",
                          data=buffer.getvalue(),
                          file_name=output_filename,
                          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                      )
                      
        except Exception as e:
            st.error(f"Error processing files: {e}")
