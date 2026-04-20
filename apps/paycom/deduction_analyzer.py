
import io
import os
import re
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st

# Use the user's active configuration file location
CONFIG_PATH = os.path.join(os.getcwd(), "Deduction Analyzer", "Deduction Setup Config.xlsx")

# st.set_page_config(page_title="Deduction Analyzer", layout="wide")


# -----------------------------
# Helpers
# -----------------------------

def norm_text(val) -> str:
    if pd.isna(val):
        return ""
    s = str(val).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def upper_clean(val) -> str:
    return norm_text(val).upper()


def yes_no(val) -> str:
    s = upper_clean(val)
    if s in {"Y", "YES", "TRUE", "1"}:
        return "Yes"
    if s in {"N", "NO", "FALSE", "0"}:
        return "No"
    return s.title() if s else ""


def is_open_ended_date(val) -> bool:
    if pd.isna(val):
        return False
    s = str(val).strip()
    if not s:
        return False
    s_nospace = s.replace(" ", "")
    zero_patterns = {
        "0000", "00/00/0000", "0/0/0000", "00-00-0000", "0-0-0000",
        "00/00/00", "0/0/00"
    }
    if s_nospace in zero_patterns:
        return True
    if re.search(r"(^|\D)0{4}(\D|$)", s_nospace):
        return True
    if s_nospace.startswith("00/") or s_nospace.endswith("/0000"):
        return True
    return False


def parse_date_safe(val) -> pd.Timestamp:
    if pd.isna(val):
        return pd.NaT
    if is_open_ended_date(val):
        return pd.NaT
    s = str(val).strip()
    if not s:
        return pd.NaT
    return pd.to_datetime(s, errors="coerce")


def classify_item(type_code: str, description: str) -> str:
    code = upper_clean(type_code)
    desc = upper_clean(description)

    tax_keywords = [
        "WITHHOLDING TAX", "SOCIAL SECURITY", "MEDICARE", "FUTA", "SUTA",
        "WORKERS COMPENSATION", "STATE W/H", "LOCAL", "SIT", "FWT",
    ]
    contribution_keywords = [" MATCH", "MEMO", "EMPLOYER MEMO", "ER MEMO"]
    deduction_keywords = [
        "MEDICAL", "DENTAL", "VISION", "401K", "ROTH", "LOAN", "AD&D",
        "STD", "VOL EE LIFE", "SUPPORT ORDER", "GARNISH", "EARNED WAGE ACCESS",
        "HEALTHCUES", "REIMBURSE", "OVERPAYMENT",
    ]

    # Per your working setup choice, keep these separate from normal deductions.
    if code in {"NSD", "PFL"} or desc in {"NEW YORK SDI", "NY PAID FAMILY LEAVE"}:
        return "Tax / Statutory Payroll Item"

    if any(k in desc for k in contribution_keywords):
        return "Contribution"

    if any(k in desc for k in tax_keywords):
        return "Tax / Statutory Payroll Item"

    if any(k in desc for k in deduction_keywords):
        return "Deduction"

    return "Review"


def classify_value_basis(group: pd.DataFrame) -> str:
    amt_non_zero = pd.to_numeric(group["Amount_num"], errors="coerce").fillna(0).ne(0).any() if "Amount_num" in group else False
    pct_non_zero = pd.to_numeric(group["Percent_num"], errors="coerce").fillna(0).ne(0).any() if "Percent_num" in group else False

    if amt_non_zero and pct_non_zero:
        return "Amount and Percent"
    if amt_non_zero:
        return "Amount"
    if pct_non_zero:
        return "Percent"
    return "No Non-Zero Amount/Percent Found"


def all_rows_zero_flag(group: pd.DataFrame) -> tuple[str, str]:
    amt_series = pd.to_numeric(group.get("Amount_num", pd.Series(dtype=float)), errors="coerce").fillna(0)
    pct_series = pd.to_numeric(group.get("Percent_num", pd.Series(dtype=float)), errors="coerce").fillna(0)
    all_zero = amt_series.eq(0).all() and pct_series.eq(0).all()
    if all_zero:
        return "Yes", f"For all {len(group)} row(s), Amount and Percent are 0.00 or blank/zero after normalization."
    return "No", "At least one row has a non-zero Amount or Percent."


def map_presence(in_sched: bool, in_prior: bool) -> str:
    if in_sched and in_prior:
        return "Present in Both"
    if in_sched:
        return "Scheduled Report Only"
    if in_prior:
        return "Prior Payroll Only"
    return "Not Found"


def setup_relevance(classification: str) -> str:
    if classification == "Deduction":
        return "Deduction Setup Needed"
    if classification == "Contribution":
        return "Contribution Setup Needed"
    if classification == "Tax / Statutory Payroll Item":
        return "Tax Handling Needed"
    return "Review"


def recommend_action(row, analysis_year: int) -> str:
    active = row.get("Active Employee Count", 0) or 0
    terminated = row.get("Terminated Employee Count", 0) or 0
    classification = row.get("Classification", "")
    running = row.get("Running Flag", "No") == "Yes"
    payroll_presence = row.get("Found in Prior Payroll", "No") == "Yes"
    zero_flag = row.get("All Rows Zero Amount/Percent?", "No") == "Yes"
    latest_stop = row.get("Latest Actual Stop Date Parsed")
    setup_needed = row.get("Setup Relevance", "")
    family_possible = row.get("Family Consolidation Possible", "")

    is_old_date = pd.notna(latest_stop) and latest_stop < pd.Timestamp(f"{analysis_year}-01-01")

    if classification == "Contribution":
        return "Setup Contribution"
    if classification == "Tax / Statutory Payroll Item":
        return "Review - Tax/Statutory"
    
    if active > 0:
        return "Setup Required"
    if payroll_presence and setup_needed == "Deduction Setup Needed":
        return "Setup Required"
        
    # Zero Amount + Old Date + Terminated
    if zero_flag and is_old_date and active == 0:
        return "Archive / Ignore"
        
    if running and active == 0 and terminated > 0:
        return "Archive / Ignore"
        
    if is_old_date and active == 0:
        return "Archive / Ignore"
        
    if zero_flag and active == 0 and terminated == 0:
        return "Archive / Ignore"
        
    if family_possible == "No":
        return "Setup Required"
        
    return "Manual Review"


def get_deduction_family(type_code: str, description: str) -> str:
    code = upper_clean(type_code)
    desc = upper_clean(description)

    if code in {"MDC", "MC1", "MC2", "MC3", "MC4"} or desc.startswith("MEDICAL"):
        return "Medical"
    if code in {"DEN", "DEN1", "DEN2", "DEN3", "DEN4"} or desc.startswith("DENTAL"):
        return "Dental"
    if code in {"VIS", "VIS1", "VIS2", "VIS3", "VIS4"} or desc.startswith("VISION"):
        return "Vision"
    if code in {"CS1", "CS2", "CS3", "CS4", "SO1", "SO2", "SO3"} or "SUPPORT ORDER" in desc:
        return "Support Order"
    if code in {"K4P", "R4P"} or "401K" in desc or "ROTH 401K" in desc:
        return "401K / Retirement"
    if code.startswith("LN") or "LOAN" in desc:
        return "Loan"
    if "GARNISH" in desc or "TAX LEVY" in desc or "WAGE ASSIGN" in desc or "CHILD SUPPORT" in desc:
        return "Garnishment / Levy"
    if "AD&D" in desc:
        return "AD&D"
    if desc == "STD":
        return "STD"
    if "VOL EE LIFE" in desc or "LIFE" in desc:
        return "Life"
    if code in {"NSD", "PFL"} or desc in {"NEW YORK SDI", "NY PAID FAMILY LEAVE"}:
        return "Statutory NY Items"
    return "Standalone / Review"


# -----------------------------
# Configuration mapping logic
# -----------------------------

def get_local_config() -> pd.DataFrame:
    try:
        if not os.path.exists(CONFIG_PATH):
            st.error(f"Config path not found: {CONFIG_PATH}")
            return pd.DataFrame()
        return pd.read_excel(CONFIG_PATH)
    except Exception as e:
        st.error(f"Error reading config: {e}")
        return pd.DataFrame()


def generate_configuration_tab(master_df: pd.DataFrame, config_df: pd.DataFrame) -> pd.DataFrame:
    if config_df.empty:
        return pd.DataFrame(columns=["Note", "Message"], data=[["Configuration File Missing", f"Please ensure the config file is placed at: {CONFIG_PATH}"]])

    # We only configure items recommended to be setup
    setup_recommendations = ["Setup Required", "Keep"]
    keeps = master_df[master_df["Recommendation"].isin(setup_recommendations)].copy()
    if keeps.empty:
        return pd.DataFrame(columns=["Note"], data=[["No 'Setup Required' recommendations found to configure."]])

    # Standardize names for matching
    config_df["_match_key"] = config_df["Company Deduction Name"].astype(str).str.strip().str.lower()
    keeps["_match_key"] = keeps["Deduction Desc"].astype(str).str.strip().str.lower()

    # Join
    merged = pd.merge(keeps, config_df, on="_match_key", how="left")

    # Dynamic Overrides: Priority is the Paycom source data for Tax and Method
    # Derived Deduction Type (Config) <- Tax Category (Paycom)
    # Deduction Method (UI) (Config) <- Value Basis (Paycom)
    
    # Map Paycom Value Basis to Uzio UI terms
    basis_to_uzio = {
        "Fixed Dollar": "Fixed $",
        "Percent": "% of Gross Pay",
        "Both": "Fixed $ / % of Gross Pay",
        "Zero/None": "Manual Check"
    }
    merged["Derived Deduction Type"] = merged["Tax Category"]
    merged["Deduction Method (UI)"] = merged["Value Basis"].map(basis_to_uzio).fillna(merged["Value Basis"])

    # We only keep the specific codes that are part of an overlap if it's a variant plan
    # Otherwise we recommend consolidation.
    
    # Select and order final Uzio columns
    uzio_cols = [
        "Company Deduction Name", "Derived Deduction Type", "Deduction Method (UI)", 
        "Amount Per Pay", "Amount %", "W-2 Box (UI)", "W-2 Label", "Sync From Benefit", 
        "Product Category Code", "Garnishment Type", "Other Garnishment Type", 
        "Auto Assign to Employee", "Weekly Schedule", "Biweekly Schedule", 
        "Semimonthly Schedule", "Arrears Applicable", "Arrears Processing Method", 
        "Flat Arrears Amount", "Assign Paycheck Limit", "Paycheck Minimum", 
        "Paycheck Maximum", "Deduction Priority", "Plan Type", "Plan ID", "Deferral Limit"
    ]
    
    # Identify missed matches
    merged["Match Status"] = np.where(merged["Company Deduction Name"].isna(), "Manual Mapping Required", "Mapped from Config")
    
    # Fill missing Company Deduction Name with Paycom Desc if match failed
    merged["Company Deduction Name"] = merged["Company Deduction Name"].fillna(merged["Deduction Desc"])

    existing_cols = [c for c in uzio_cols if c in merged.columns]
    final_uzio = merged[["Match Status"] + existing_cols].copy()
    
    return final_uzio


# -----------------------------
# Scheduled deduction processing
# -----------------------------

def build_scheduled_summary(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    required_cols = [
        "EE Code", "EE Status", "Deduction Code", "Deduction Desc", "Memo Only",
        "Amount", "Percent", "Tax Treatment", "Start Date", "Stop Date"
    ]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Scheduled deduction file is missing columns: {missing}")

    work = df.copy()
    work["EE Code"] = work["EE Code"].astype(str).str.strip()
    work["EE Status"] = work["EE Status"].astype(str).str.strip().str.upper()
    work["Deduction Code"] = work["Deduction Code"].map(norm_text)
    work["Deduction Desc"] = work["Deduction Desc"].map(norm_text)
    work["Memo Only Clean"] = work["Memo Only"].map(yes_no)
    work["Amount_num"] = pd.to_numeric(work["Amount"], errors="coerce").fillna(0)
    work["Percent_num"] = pd.to_numeric(work["Percent"], errors="coerce").fillna(0)
    work["Start Date Text"] = work["Start Date"].astype(str).str.strip()
    work["Stop Date Text"] = work["Stop Date"].astype(str).str.strip()
    work["Start Open"] = work["Start Date"].apply(is_open_ended_date)
    work["Stop Open"] = work["Stop Date"].apply(is_open_ended_date)
    work["Open Date Flag"] = work["Start Open"] | work["Stop Open"]
    work["Start Date Parsed"] = work["Start Date"].apply(parse_date_safe)
    work["Stop Date Parsed"] = work["Stop Date"].apply(parse_date_safe)
    work["Deduction Family"] = work.apply(lambda r: get_deduction_family(r["Deduction Code"], r["Deduction Desc"]), axis=1)

    group_cols = ["Deduction Code", "Deduction Desc"]

    def summarize_group(g: pd.DataFrame) -> pd.Series:
        running = g["Open Date Flag"].any()
        actual_rows = g.loc[~g["Open Date Flag"]].copy()
        active_ee = g.loc[g["EE Status"] == "A", "EE Code"].nunique()
        term_ee = g.loc[g["EE Status"] == "T", "EE Code"].nunique()
        v_ee = g.loc[g["EE Status"] == "V", "EE Code"].nunique()
        value_basis_raw = classify_value_basis(g)
        basis_map = {
            "Amount": "Fixed Dollar",
            "Percent": "Percent",
            "Amount and Percent": "Both",
            "No Non-Zero Amount/Percent Found": "Zero/None"
        }
        value_basis = basis_map.get(value_basis_raw, value_basis_raw)

        tax_raw = g["Tax Treatment"].dropna().unique().tolist()
        tax_val = tax_raw[0] if tax_raw else ""
        
        if "After Tax" in tax_val:
            tax_cat = "Post-tax"
        elif "Pre-Tax" in tax_val or "Taxable Only (401k)" in tax_val:
            tax_cat = "Pre-tax"
        else:
            tax_cat = "Review"

        zero_flag, zero_note = all_rows_zero_flag(g)

        if running:
            start_display = "00/00/0000"
            stop_display = "00/00/0000"
        else:
            start_candidates = actual_rows["Start Date Parsed"].dropna()
            stop_candidates = actual_rows["Stop Date Parsed"].dropna()
            start_display = start_candidates.min().strftime("%m/%d/%Y") if not start_candidates.empty else ""
            stop_display = stop_candidates.max().strftime("%m/%d/%Y") if not stop_candidates.empty else ""

        memo_vals = [x for x in g["Memo Only Clean"].dropna().unique().tolist() if x != ""]
        memo_out = ", ".join(sorted(memo_vals)) if memo_vals else ""

        latest_stop_parsed = actual_rows["Stop Date Parsed"].dropna().max() if not actual_rows.empty else pd.NaT
        earliest_start_parsed = actual_rows["Start Date Parsed"].dropna().min() if not actual_rows.empty else pd.NaT
        family = g["Deduction Family"].mode().iloc[0] if not g["Deduction Family"].dropna().empty else "Standalone / Review"

        return pd.Series({
            "Deduction Family": family,
            "Memo Only": memo_out,
            "Running Flag": "Yes" if running else "No",
            "Date Classification": "Running Deduction" if running else "Actual Dated Deduction",
            "Start Date": start_display,
            "Stop Date": stop_display,
            "Earliest Actual Start Date Parsed": earliest_start_parsed,
            "Latest Actual Stop Date Parsed": latest_stop_parsed,
            "Scheduled Row Count": len(g),
            "Scheduled Distinct Employees": g["EE Code"].nunique(),
            "Active Employee Count": active_ee,
            "Terminated Employee Count": term_ee,
            "V Status Employee Count": v_ee,
            "Amount or Percent": value_basis_raw,
            "Value Basis": value_basis,
            "Tax Treatment": tax_val,
            "Tax Category": tax_cat,
            "All Rows Zero Amount/Percent?": zero_flag,
            "Zero Value Note": zero_note,
        })

    summary = work.groupby(group_cols, dropna=False).apply(summarize_group).reset_index()
    return summary, work


# -----------------------------
# Prior payroll processing
# -----------------------------

def build_prior_summary(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    required_cols = ["EE Code", "Type Code", "Type Description", "Amount"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        raise ValueError(f"Prior payroll file is missing columns: {missing}")

    work = df.copy()
    work["EE Code"] = work["EE Code"].astype(str).str.strip()
    work["Type Code"] = work["Type Code"].map(norm_text)
    work["Type Description"] = work["Type Description"].map(norm_text)
    work["Amount_num"] = pd.to_numeric(work["Amount"], errors="coerce").fillna(0)
    work["Deduction Family"] = work.apply(lambda r: get_deduction_family(r["Type Code"], r["Type Description"]), axis=1)

    if "Code Description" in work.columns:
        work["Code Description"] = work["Code Description"].fillna("").astype(str)
        filtered = work.loc[work["Code Description"].str.strip().str.upper() == "DEDUCTIONS"].copy()
    else:
        excluded_desc = {
            "REGULAR", "BONUS (HOURS)", "PTO", "PAID TIME OFF", "NET CHECK",
            "NET DISTRIBUTION 1", "NET DISTRIBUTION 2", "RETRO REGULAR PAY"
        }
        filtered = work.loc[~work["Type Description"].str.upper().isin(excluded_desc)].copy()

    summary = (
        filtered.groupby(["Type Code", "Type Description"], dropna=False)
        .agg(
            Prior_Payroll_Row_Count=("EE Code", "size"),
            Prior_Payroll_Distinct_Employees=("EE Code", "nunique"),
            Prior_Payroll_Total_Amount=("Amount_num", "sum"),
            Prior_Payroll_NonZero_Row_Count=("Amount_num", lambda s: int(s.ne(0).sum())),
            Prior_Deduction_Family=("Deduction Family", lambda s: s.mode().iloc[0] if not s.mode().empty else "Standalone / Review"),
        )
        .reset_index()
        .rename(columns={"Type Code": "Deduction Code", "Type Description": "Deduction Desc"})
    )
    return summary, filtered


# -----------------------------
# Family overlap analysis
# -----------------------------

def build_family_analysis(scheduled_detail: pd.DataFrame) -> pd.DataFrame:
    base = scheduled_detail.copy()
    base["_Classification"] = base.apply(lambda r: classify_item(r.get("Deduction Code", ""), r.get("Deduction Desc", "")), axis=1)
    base = base[base["_Classification"] != "Contribution"].copy()
    
    base["EE Code"] = base["EE Code"].astype(str).str.strip()
    base["Deduction Code"] = base["Deduction Code"].map(norm_text)
    base["Deduction Desc"] = base["Deduction Desc"].map(norm_text)
    base["Deduction Family"] = base["Deduction Family"].fillna("Standalone / Review")

    # One employee can have many rows for the same code; dedupe at employee+family+code level.
    emp_code = (
        base[["EE Code", "EE Name", "Deduction Family", "Deduction Code", "Deduction Desc"]]
        .drop_duplicates()
    )

    family_totals = (
        emp_code.groupby("Deduction Family")
        .agg(
            Family_Distinct_Employees=("EE Code", "nunique"),
            Family_Distinct_Codes=("Deduction Code", "nunique"),
        )
        .reset_index()
    )

    emp_family_counts = (
        emp_code.groupby(["Deduction Family", "EE Code"])
        .agg(
            EE_Name=("EE Name", "first"),
            Employee_Family_Code_Count=("Deduction Code", "nunique"),
            Employee_Family_Codes=("Deduction Code", lambda s: list(set(s))),
            Employee_Family_Code_Str=("Deduction Code", lambda s: ", ".join(sorted(set(s)))),
        )
        .reset_index()
    )

    overlap = emp_family_counts.loc[emp_family_counts["Employee_Family_Code_Count"] > 1].copy()

    # Identify which SPECIFIC codes are part of overlaps
    all_overlapping_codes_per_family = {}
    family_overlap_names = {}
    
    for _, row in overlap.iterrows():
        fam = row["Deduction Family"]
        codes = row["Employee_Family_Codes"]
        name = row["EE_Name"]
        if fam not in all_overlapping_codes_per_family:
            all_overlapping_codes_per_family[fam] = set()
            family_overlap_names[fam] = set()
        all_overlapping_codes_per_family[fam].update(codes)
        family_overlap_names[fam].add(name)

    overlap_summary = (
        overlap.groupby("Deduction Family")
        .agg(
            Family_Overlap_Employee_Count=("EE Code", "nunique"),
            Overlap_Employee_IDs=("EE Code", lambda s: ", ".join(sorted(set(map(str, s))))),
            Overlap_Code_Combos=("Employee_Family_Code_Str", lambda s: " | ".join(sorted(set(s)))),
        )
        .reset_index()
    )

    member_codes = (
        emp_code.groupby("Deduction Family")
        .agg(
            Family_Member_Codes=("Deduction Code", lambda s: ", ".join(sorted(set(s)))),
            Family_Member_Descriptions=("Deduction Desc", lambda s: " | ".join(sorted(set(s)))),
        )
        .reset_index()
    )

    family = family_totals.merge(member_codes, on="Deduction Family", how="left")
    family = family.merge(overlap_summary, on="Deduction Family", how="left")

    family["Family_Overlap_Employee_Count"] = family["Family_Overlap_Employee_Count"].fillna(0).astype(int)
    family["Overlap Exists"] = np.where(family["Family_Overlap_Employee_Count"] > 0, "Yes", "No")
    
    # Store the overlapping codes in the family analysis for later lookup
    family["Overlapping_Codes_Set"] = family["Deduction Family"].map(lambda x: tuple(sorted(list(all_overlapping_codes_per_family.get(x, [])))))
    family["Overlapping_Employee_Names"] = family["Deduction Family"].map(lambda x: ", ".join(sorted(list(family_overlap_names.get(x, []))))[:200])

    family["Family Consolidation Possible"] = np.where(family["Family_Overlap_Employee_Count"] > 0, "Partial", "Yes")
    family["Family Recommended Setup"] = np.where(
        family["Family_Overlap_Employee_Count"] > 0,
        "Keep Overlapping Variants, Merge Others",
        "Can Consolidate All into Single Plan"
    )
    family["Overlap_Employee_IDs"] = family["Overlap_Employee_IDs"].fillna("")
    family["Overlap_Code_Combos"] = family["Overlap_Code_Combos"].fillna("")
    return family.sort_values(["Family_Overlap_Employee_Count", "Deduction Family"], ascending=[False, True]).reset_index(drop=True)


# -----------------------------
# Merge / final workbook
# -----------------------------

def build_master_sheet(
    scheduled_summary: pd.DataFrame,
    prior_summary: pd.DataFrame,
    family_analysis: pd.DataFrame,
    analysis_year: int
) -> pd.DataFrame:
    master = pd.merge(
        scheduled_summary,
        prior_summary,
        on=["Deduction Code", "Deduction Desc"],
        how="outer",
    )

    for col in [
        "Scheduled Row Count", "Scheduled Distinct Employees", "Active Employee Count",
        "Terminated Employee Count", "V Status Employee Count", "Prior_Payroll_Row_Count",
        "Prior_Payroll_Distinct_Employees", "Prior_Payroll_Total_Amount", "Prior_Payroll_NonZero_Row_Count",
    ]:
        if col in master.columns:
            master[col] = master[col].fillna(0)

    master["Found in Scheduled Report"] = np.where(master["Scheduled Row Count"].fillna(0).gt(0), "Yes", "No")
    master["Found in Prior Payroll"] = np.where(master["Prior_Payroll_Row_Count"].fillna(0).gt(0), "Yes", "No")
    master["Source Presence"] = master.apply(
        lambda r: map_presence(r["Found in Scheduled Report"] == "Yes", r["Found in Prior Payroll"] == "Yes"),
        axis=1,
    )

    for col in [
        "Memo Only", "Running Flag", "Date Classification", "Start Date", "Stop Date",
        "Amount or Percent", "All Rows Zero Amount/Percent?", "Zero Value Note",
        "Deduction Family"
    ]:
        if col not in master.columns:
            master[col] = ""
        master[col] = master[col].fillna("")

    if "Prior_Deduction_Family" in master.columns:
        master["Deduction Family"] = np.where(
            master["Deduction Family"].astype(str).str.strip().eq(""),
            master["Prior_Deduction_Family"].fillna(""),
            master["Deduction Family"]
        )

    master["Classification"] = master.apply(
        lambda r: classify_item(r.get("Deduction Code", ""), r.get("Deduction Desc", "")),
        axis=1,
    )
    master["Setup Relevance"] = master["Classification"].map(setup_relevance)

    family_cols = [
        "Deduction Family",
        "Family_Distinct_Employees",
        "Family_Distinct_Codes",
        "Family_Member_Codes",
        "Family_Overlap_Employee_Count",
        "Overlap Exists",
        "Overlap_Employee_IDs",
        "Overlap_Code_Combos",
        "Family Consolidation Possible",
        "Family Recommended Setup",
        "Overlapping_Codes_Set",
        "Overlapping_Employee_Names"
    ]
    family_merge = family_analysis[family_cols].drop_duplicates()
    master = master.merge(family_merge, on="Deduction Family", how="left")

    for col in [
        "Family_Distinct_Employees", "Family_Distinct_Codes", "Family_Overlap_Employee_Count"
    ]:
        if col in master.columns:
            master[col] = master[col].fillna(0).astype(int)

    for col in [
        "Overlap Exists", "Overlap_Employee_IDs", "Overlap_Code_Combos",
        "Family Consolidation Possible", "Family Recommended Setup"
    ]:
        if col in master.columns:
            master[col] = master[col].fillna("")

    master["Recommended Setup Structure"] = np.where(
        master["Family Consolidation Possible"] == "No",
        "Keep Separate",
        np.where(
            master["Deduction Family"].isin(["Medical", "Dental", "Vision", "Support Order", "401K / Retirement", "Loan", "Garnishment / Levy"]),
            "Can Consolidate",
            "Review"
        )
    )

    master["Recommendation"] = master.apply(lambda r: recommend_action(r, analysis_year), axis=1)

    # Apply Overlap-Only Consolidation Correction
    def refine_overlap_recommendation(row):
        rec = row["Recommendation"]
        if "Setup Required" not in rec:
            return rec
            
        fam = row["Deduction Family"]
        code = row["Deduction Code"]
        desc = row["Deduction Desc"]
        
        # Identify Standard/Primary Plan
        primary_keywords = ["STANDARD", "MEDICAL", "DENTAL", "VISION", "SUPPORT ORDER", "BASIC"]
        is_primary = any(k in str(desc).upper() for k in primary_keywords) and not re.search(r"\d", str(desc))
        
        # Get overlapping codes for this family
        over_codes = row.get("Overlapping_Codes_Set", [])
        if not isinstance(over_codes, (list, tuple)):
            over_codes = []
        
        if code in over_codes:
            return "Setup Required (Overlapping)"
        
        if is_primary:
            return "Setup Required (Standard)"
            
        # Any other variant plan that isn't overlapping should be avoided/merged
        if "Medical" in fam or "Dental" in fam or "Vision" in fam or "Support Order" in fam or "401k" in fam.lower():
            return "Avoid - Consolidate into Standard"
            
        return rec

    master["Recommendation"] = master.apply(refine_overlap_recommendation, axis=1)

    master["Review Note"] = ""
    master.loc[
        (master["Found in Scheduled Report"] == "No") & (master["Found in Prior Payroll"] == "Yes"),
        "Review Note",
    ] = "Present in prior payroll but not found in scheduled deduction report."
    master.loc[
        (master["Found in Scheduled Report"] == "Yes") & (master["Found in Prior Payroll"] == "No"),
        "Review Note",
    ] = "Present in scheduled deduction report but not used in this prior payroll file/pay period."
    master.loc[
        master["All Rows Zero Amount/Percent?"] == "Yes",
        "Review Note",
    ] = master["Review Note"].astype(str).str.strip() + " All scheduled rows have zero Amount/Percent."

    master.loc[
        master["Family Consolidation Possible"] == "No",
        "Review Note",
    ] = (
        master["Review Note"].astype(str).str.strip()
        + " Family overlap exists, so keep separate family member deductions."
    ).str.strip()

    ordered_cols = [
        "Deduction Code", "Deduction Desc", "Deduction Family", "Classification", "Setup Relevance",
        "Tax Category", "Value Basis", "Recommended Setup Structure", "Recommendation",
        "Source Presence", "Found in Scheduled Report", "Found in Prior Payroll",
        "Running Flag", "Date Classification", "Start Date", "Stop Date", "Tax Treatment", "Memo Only",
        "Scheduled Row Count", "Scheduled Distinct Employees", "Active Employee Count",
        "Terminated Employee Count", "V Status Employee Count",
        "Amount or Percent", "All Rows Zero Amount/Percent?", "Zero Value Note",
        "Prior_Payroll_Row_Count", "Prior_Payroll_Distinct_Employees",
        "Prior_Payroll_Total_Amount", "Prior_Payroll_NonZero_Row_Count",
        "Family_Distinct_Employees", "Family_Distinct_Codes", "Family_Member_Codes",
        "Family_Overlap_Employee_Count", "Overlap Exists", "Overlap_Employee_IDs",
        "Overlap_Code_Combos", "Family Consolidation Possible", "Family Recommended Setup",
        "Family Rule Note", "Review Note",
        "Earliest Actual Start Date Parsed", "Latest Actual Stop Date Parsed",
    ]
    existing = [c for c in ordered_cols if c in master.columns]
    remaining = [c for c in master.columns if c not in existing]
    master = master[existing + remaining].sort_values(
        ["Recommendation", "Deduction Family", "Deduction Code", "Deduction Desc"]
    ).reset_index(drop=True)
    return master


def to_excel_bytes(
    master: pd.DataFrame,
    family_analysis: pd.DataFrame,
    scheduled_detail: pd.DataFrame,
    prior_detail: pd.DataFrame,
    config_guide: pd.DataFrame = None
) -> bytes:
    output = io.BytesIO()
    
    # 1. Prepare Action Checklist
    action_cols = [
        "Deduction Code", "Deduction Desc", "Recommendation", 
        "Classification", "Source Presence", "Family Recommended Setup"
    ]
    action_df = master[[c for c in action_cols if c in master.columns]].copy()
    
    sort_order = {
        "Setup Required (Overlapping)": 0,
        "Setup Required (Standard)": 1,
        "Setup Required": 2, 
        "Setup Contribution": 3, 
        "Manual Review": 4, 
        "Avoid - Consolidate into Standard": 5,
        "Review - Tax/Statutory": 6, 
        "Archive / Ignore": 7
    }
    action_df["_sort"] = action_df["Recommendation"].map(sort_order).fillna(8)
    action_df = action_df.sort_values(by=["_sort", "Recommendation", "Deduction Desc"]).drop(columns=["_sort"])
    
    # 2. Prepare Consolidation Plan
    consol_cols = ["Deduction Family", "Deduction Code", "Deduction Desc", "Recommendation", "Overlapping_Employee_Names"]
    if "Overlapping_Employee_Names" in master.columns:
        consol_df = master[master["Recommendation"].isin(["Setup Required (Overlapping)", "Avoid - Consolidate into Standard"])][consol_cols].copy()
        consol_df["Reasoning"] = np.where(
            consol_df["Recommendation"] == "Setup Required (Overlapping)",
            "Conflict found: Employee(s) [" + consol_df["Overlapping_Employee_Names"] + "] have multiple plans. Must keep separate.",
            "No employee-level overlaps found. Safe to merge into Standard plan to simplify setup."
        )
        consol_df = consol_df.sort_values(["Deduction Family", "Recommendation"])
    else:
        consol_df = pd.DataFrame(columns=["Note"], data=[["No consolidation logic required for this file."]])

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        action_df.to_excel(writer, index=False, sheet_name="Action_Checklist")
        if config_guide is not None and not config_guide.empty:
            config_guide.to_excel(writer, index=False, sheet_name="Uzio_Setup_Guide")
        if not consol_df.empty:
            consol_df.to_excel(writer, index=False, sheet_name="Consolidation_Plan")
            
        wb = writer.book
        from openpyxl.styles import PatternFill, Font
        blue_fill = PatternFill("solid", fgColor="D9EAF7")
        green_fill = PatternFill("solid", fgColor="E2F0D9")
        yellow_fill = PatternFill("solid", fgColor="FFF2CC")
        red_fill = PatternFill("solid", fgColor="F4CCCC")
        purple_fill = PatternFill("solid", fgColor="E2D9F7")
        bold = Font(bold=True)

        def adjust_cols(ws):
            for col_cells in ws.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells[:2000]:
                    try:
                        max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
                    except: pass
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)

        # 1. Format Action_Checklist
        if "Action_Checklist" in writer.sheets:
            ws_ac = writer.sheets["Action_Checklist"]
            ws_ac.freeze_panes = "A2"
            for cell in ws_ac[1]:
                cell.font = bold
                cell.fill = blue_fill
            adjust_cols(ws_ac)
            
            headers = [c.value for c in ws_ac[1]]
            idx = {name: i + 1 for i, name in enumerate(headers)}
            for row in range(2, ws_ac.max_row + 1):
                rec = ws_ac.cell(row, idx.get("Recommendation", 1)).value if idx.get("Recommendation") else ""
                fill = None
                if rec == "Archive / Ignore":
                    fill = red_fill
                elif rec == "Manual Review" or "Review" in str(rec):
                    fill = yellow_fill
                elif "Avoid" in str(rec):
                    fill = purple_fill
                elif "Setup" in str(rec):
                    fill = green_fill
                if fill:
                    for col in range(1, ws_ac.max_column + 1):
                        ws_ac.cell(row, col).fill = fill

        # 2. Format Uzio_Setup_Guide
        if "Uzio_Setup_Guide" in writer.sheets:
            ws2 = writer.sheets["Uzio_Setup_Guide"]
            ws2.freeze_panes = "B2"
            for cell in ws2[1]:
                cell.font = bold
                cell.fill = blue_fill
            adjust_cols(ws2)

        # 3. Format Consolidation_Plan
        if "Consolidation_Plan" in writer.sheets:
            ws_cp = writer.sheets["Consolidation_Plan"]
            ws_cp.freeze_panes = "A2"
            for cell in ws_cp[1]:
                cell.font = bold
                cell.fill = blue_fill
            adjust_cols(ws_cp)
            
            headers = [c.value for c in ws_cp[1]]
            idx = {name: i + 1 for i, name in enumerate(headers)}
            for row in range(2, ws_cp.max_row + 1):
                rec = ws_cp.cell(row, idx.get("Recommendation", 1)).value if idx.get("Recommendation") else ""
                fill = green_fill if "Overlapping" in str(rec) else purple_fill
                for col in range(1, ws_cp.max_column + 1):
                    ws_cp.cell(row, col).fill = fill

    output.seek(0)
    return output.getvalue()


# -----------------------------
# UI
# -----------------------------

def render_ui():
    st.title("Scheduled Deduction + Prior Payroll Analyzer")
    st.write(
        "Upload a Paycom **Scheduled Deduction Report** and a **Prior Payroll** file. "
        "The app will create one smart summary sheet plus family overlap analysis to help decide "
        "what should be set up as company-level deductions in UZIO."
    )

    col1, col2, col3 = st.columns([1, 1, 1])
    with col1:
        scheduled_file = st.file_uploader("Scheduled Deduction Report", type=["xlsx", "csv"], key="scheduled")
    with col2:
        prior_file = st.file_uploader("Prior Payroll File", type=["xlsx", "csv"], key="prior")
    with col3:
        analysis_year = st.number_input("Analysis year for end-date checks", min_value=2020, max_value=2035, value=2026, step=1)

    run = st.button("Analyze Files", type="primary", use_container_width=True)

    if run:
        if not scheduled_file or not prior_file:
            st.error("Please upload both files first.")
            st.stop()

        try:
            def load_df(file):
                if file.name.lower().endswith(".csv"):
                    try:
                        return pd.read_csv(file)
                    except UnicodeDecodeError:
                        return pd.read_csv(file, encoding="latin1")
                return pd.read_excel(file)

            sched_df = load_df(scheduled_file)
            prior_df = load_df(prior_file)

            scheduled_summary, scheduled_detail = build_scheduled_summary(sched_df)
            prior_summary, prior_detail = build_prior_summary(prior_df)
            family_analysis = build_family_analysis(scheduled_detail)
            master = build_master_sheet(
                scheduled_summary=scheduled_summary,
                prior_summary=prior_summary,
                family_analysis=family_analysis,
                analysis_year=int(analysis_year)
            )
            
            # Persistent Config Integration
            config_df = get_local_config()
            config_guide = generate_configuration_tab(master, config_df)
            
            excel_bytes = to_excel_bytes(master, family_analysis, scheduled_detail, prior_detail, config_guide)

            st.success("Analysis completed.")
            if not config_df.empty:
                mapped_count = (config_guide["Match Status"] == "Mapped from Config").sum()
                st.info(f"Matched {mapped_count} deductions to standard Uzio configurations.")
            else:
                st.warning("Local configuration file (deduction_setup_config.xlsx) not found. Setup guide will be generic.")

            k1, k2, k3, k4 = st.columns(4)
            k1.metric("Total items in master", len(master))
            k2.metric("Keep", int((master["Recommendation"] == "Keep").sum()))
            k3.metric("Remove candidate", int((master["Recommendation"] == "Remove Candidate").sum()))
            k4.metric("Families with overlap", int((family_analysis["Family_Overlap_Employee_Count"] > 0).sum()))

            st.subheader("Master Summary Preview")
            st.dataframe(master, use_container_width=True, height=500)
            
            if not config_guide.empty:
                st.subheader("Uzio Configuration Preview")
                st.dataframe(config_guide, use_container_width=True, height=400)

            st.subheader("Family Analysis Preview")
            st.dataframe(family_analysis, use_container_width=True, height=300)

            st.download_button(
                "Download Smart Analysis Workbook",
                data=excel_bytes,
                file_name=f"deduction_smart_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

            with st.expander("Business rules used"):
                st.markdown(
                    """
- **Dual-Tab Output**: Produces a summary recommendation tab and a technical configuration tab.
- **Tax & Method**: These are derived from Paycom's source files to ensure accuracy.
- **Standard Mapping**: uzio-specific settings (W-2, Arrears, etc.) are looked up from the saved `deduction_setup_config.xlsx`.
- If **any row** of a scheduled deduction has `0000` in Start Date or Stop Date, it is treated as a **Running Deduction**.
- `A`, `T`, and `V` employee statuses are counted separately from the scheduled deduction report.
- Deduction families such as **Medical** and **Support Order** are analyzed at employee level.
- If any employee has more than one code within the same family, then **Family Consolidation Possible = No**.
                    """
                )

        except Exception as e:
            st.exception(e)


if __name__ == "__main__":
    render_ui()
