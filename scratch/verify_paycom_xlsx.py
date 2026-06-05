"""End-to-end check: extract earnings/deductions/contributions from the
Chief Delivery prior payroll file, build the 3-tab xlsx, write it to disk,
re-read each tab, print contents."""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
from apps.paycom.prior_payroll_setup_helper import (
    extract_unique_earnings_from_prior,
    extract_unique_deductions_from_prior,
    bifurcate_match_memo,
    classify_pre_post_from_calc_description,
    build_3tab_setup_xlsx,
)

PATH = r"C:\Users\rohit.kaushik\Downloads\Chief Delivery\payroll\PriorPayroll_12212025_03142026_03202026.csv"
OUT  = r"C:\Users\rohit.kaushik\Downloads\Chief Delivery\REBUILD_Chief_Delivery_Setup_Helper.xlsx"

df = pd.read_csv(PATH, dtype=str)
earnings = extract_unique_earnings_from_prior(df)
all_deds = extract_unique_deductions_from_prior(df)
contribs, deds = bifurcate_match_memo(all_deds)
deds = classify_pre_post_from_calc_description(df, deds)

data = build_3tab_setup_xlsx(earnings=earnings, deductions=deds, contributions=contribs)
with open(OUT, "wb") as f:
    f.write(data)
print(f"Wrote {OUT} ({len(data):,} bytes)\n")

# Re-read each tab to verify
from openpyxl import load_workbook
wb = load_workbook(OUT, read_only=True, data_only=True)
print(f"Sheets in the file: {wb.sheetnames}\n")
for n in wb.sheetnames:
    ws = wb[n]
    print(f"=== Tab: {n!r}  ({ws.max_row} rows x {ws.max_column} cols) ===")
    sub = pd.read_excel(OUT, sheet_name=n, dtype=str)
    print(sub.to_string(index=False))
    print()
wb.close()
