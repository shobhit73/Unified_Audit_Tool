import io
import re
import pandas as pd
import streamlit as st
import openpyxl
import difflib
import numpy as np
from collections import defaultdict
from datetime import datetime

APP_TITLE = "ADP → Uzio Prior Payroll Generator"

# ─── Uzio template constants ───
UZIO_HEADER_ROW = 5
UZIO_SECTION_ROW = 4
UZIO_DATA_START_ROW = 6
UZIO_EMPLOYEE_ID_COL = 1
UZIO_FULL_NAME_COL = 2
UZIO_SSN_COL = 3
UZIO_PP_START_COL = 4
UZIO_PP_END_COL = 5
UZIO_PAYCHECK_DATE_COL = 6
UZIO_FIRST_DATA_COL = 7

UI_SECTIONS_ORDER = ['Earnings', 'Deductions', 'Contributions', 'Employee Taxes', 'Employer Taxes']
SKIP_LABEL = "── Skip (do not map) ──"

# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def get_adp_category(col_name: str) -> str:
    """Categorize an ADP dynamic column into a UI section."""
    name = str(col_name).strip().upper()
    
    # Skip items
    if ('MEMO' in name or 
        'DIRECT DEPOSIT' in name or 
        name in ('GROSS PAY', 'TAKE HOME', 'NET PAY') or 
        name.startswith('TOTAL ')):
        return '_SKIP'
        
    if 'EARNINGS' in name or 'HOURS' in name:
        return 'Earnings'
    if 'EMPLOYEE TAX' in name:
        return 'Employee Taxes'
    if 'EMPLOYER TAX' in name:
        return 'Employer Taxes'
    if 'DEDUCTION' in name:
        return 'Deductions'
        
    return 'Deductions' # Fallback


def auto_guess_mapping(adp_col: str, uzio_col_headers: dict) -> str:
    """Attempt to auto-map ADP column string to Uzio column string."""
    if not adp_col or not isinstance(adp_col, str):
        return SKIP_LABEL
        
    td_lower = adp_col.lower()
    td_clean = re.sub(r'[^a-z0-9]', '', td_lower)
    if not td_clean:
        return SKIP_LABEL
        
    best_match = SKIP_LABEL
    best_score = 0.0
    
    for col_idx, hdr in uzio_col_headers.items():
        if col_idx < UZIO_FIRST_DATA_COL:
            continue
            
        hdr_lower = str(hdr).lower()
        hdr_clean = re.sub(r'[^a-z0-9]', '', hdr_lower)
        if not hdr_clean:
            continue
            
        score = difflib.SequenceMatcher(None, td_clean, hdr_clean).ratio()
        
        td_words = set(re.findall(r'[a-z0-9]+', td_lower))
        hdr_words = set(re.findall(r'[a-z0-9]+', hdr_lower))
        overlap = td_words & hdr_words
        if overlap:
            score += 0.15 * len(overlap)
            
        # Domain boosts
        if 'medicare' in td_words and 'medicare' in hdr_words: score += 0.3
        if ('soc' in td_words or 'ss' in td_words) and 'social' in hdr_words: score += 0.3
        if ('fit' in td_words or 'fed' in td_words or 'federal' in td_words) and 'federal' in hdr_words: score += 0.3
        if '401k' in td_lower and '401k' in hdr_lower: score += 0.3
        if 'regular' in td_words and 'regular' in hdr_words: score += 0.3
        if 'overtime' in td_words and 'overtime' in hdr_words: score += 0.3
        if 'bonus' in td_words and 'bonus' in hdr_words: score += 0.3
        if 'futa' in td_words and 'futa' in hdr_words: score += 0.3
        if 'sui' in td_words and 'sui' in hdr_words: score += 0.3
        if 'sdi' in td_words and 'sdi' in hdr_words: score += 0.3
        if 'worked in state' in td_lower and 'state income' in hdr_lower: score += 0.5
        
        if score > best_score and score >= 0.65:
            best_score = score
            best_match = f"Col {col_idx}: {hdr}"
            
    return best_match


def reformat_name(raw_name: str) -> str:
    """Ensure name is 'LAST, FIRST'. Note: ADP already typically uses Last, First, so just clean it up."""
    if not raw_name or not isinstance(raw_name, str):
        return raw_name or ""
    return raw_name.strip()


def parse_date(date_str):
    if pd.isna(date_str) or not str(date_str).strip():
        return ""
    try:
        if isinstance(date_str, datetime):
            return date_str.strftime('%m/%d/%Y')
        pt = pd.to_datetime(date_str)
        return pt.strftime('%m/%d/%Y')
    except:
        return str(date_str).strip()


def read_uzio_template(file_bytes):
    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    ws = wb[wb.sheetnames[-1]]
    
    for sname in wb.sheetnames:
        if 'payroll' in sname.lower() and 'instruction' not in sname.lower():
            ws = wb[sname]
            break
            
    section_headers = {}
    for cell in ws[UZIO_SECTION_ROW]:
        if cell.value:
            section_headers[cell.column] = str(cell.value).strip()
            
    column_headers = {}
    for cell in ws[UZIO_HEADER_ROW]:
        if cell.value:
            column_headers[cell.column] = str(cell.value).strip()
            
    return section_headers, column_headers, wb, ws


def read_adp_files(uploaded_files):
    """Read ADP files and extract unique dynamic columns."""
    all_rows = []
    all_dynamic_cols = set()
    
    file_summaries = []
    
    for f in uploaded_files:
        try:
            xl = pd.ExcelFile(io.BytesIO(f.read()))
            f.seek(0)
            sheet = xl.sheet_names[0]
            
            # Read header
            df_test = xl.parse(sheet, nrows=10)
            
            # Find row with FILE NUMBER
            header_row_idx = None
            for i in range(len(df_test)):
                row_vals = [str(x).upper().strip() for x in df_test.iloc[i].tolist()]
                if 'FILE NUMBER' in row_vals or 'COMPANY CODE' in row_vals:
                    header_row_idx = i
                    break
                    
            if header_row_idx is not None:
                df = xl.parse(sheet, header=header_row_idx+1)
            else:
                df = xl.parse(sheet)
                
            orig_len = len(df)
                
            # Filter valid rows (exclude totals, blanks)
            if 'FILE NUMBER' in df.columns:
                df = df[df['FILE NUMBER'].notna()]
                df = df[~df['FILE NUMBER'].astype(str).str.contains('Total', case=False, na=False)]
            elif 'COMPANY CODE' in df.columns:
                df = df[df['COMPANY CODE'].notna()]
                df = df[~df['COMPANY CODE'].astype(str).str.contains('Total', case=False, na=False)]
                
            if 'NAME' in df.columns:
                df = df[df['NAME'].notna()]
            
            # Discover dynamic columns (anything after standard identifiers)
            standard_cols = ['COMPANY CODE', 'NAME', 'FILE NUMBER', 'POSITION ID', 'STATUS', 'TAX ID', 'ASSOCIATE ID', 'WORKED IN STATE', 'DIST #', 'PERIOD BEGINNING DATE', 'PERIOD ENDING DATE', 'PAY DATE', 'CHECK/VOUCHER NUMBER']
            
            file_dynamic_cols = [c for c in df.columns if c not in standard_cols and not str(c).startswith('Unnamed:')]
            all_dynamic_cols.update(file_dynamic_cols)
            
            if not df.empty:
                min_start = df['PERIOD BEGINNING DATE'].min() if 'PERIOD BEGINNING DATE' in df.columns else ""
                max_end = df['PERIOD ENDING DATE'].max() if 'PERIOD ENDING DATE' in df.columns else ""
                pay_date = df['PAY DATE'].iloc[0] if 'PAY DATE' in df.columns else ""
                
                file_summaries.append({
                    'Filename': f.name,
                    'Pay Period': f"{parse_date(min_start)} - {parse_date(max_end)}",
                    'Pay Date': parse_date(pay_date),
                    'Employees': len(df['FILE NUMBER'].unique()) if 'FILE NUMBER' in df.columns else 0,
                    'Records': len(df)
                })
            
            # Store data
            for _, row in df.iterrows():
                # Clean up NaN to None
                row_dict = {k: (v if pd.notna(v) else None) for k, v in row.to_dict().items()}
                all_rows.append(row_dict)
                
        except Exception as e:
            st.error(f"Failed to process {f.name}: {e}")
            
    return all_rows, list(all_dynamic_cols), file_summaries


def generate_output(adp_rows, mapping, uzio_col_headers, net_pay_col_idx):
    output_rows = []
    skipped_items = []
    validation_results = []
    
    # Standardize column mappings so we can check easily
    # mapping is { adp_col_name : target_uzio_col_idx }
    
    # Group by EE, Pay Period Start
    ee_groups = defaultdict(list)
    for r in adp_rows:
        ee_code = str(r.get('FILE NUMBER', r.get('ASSOCIATE ID', ''))).strip()
        pp_start = parse_date(r.get('PERIOD BEGINNING DATE'))
        if ee_code:
            ee_groups[(ee_code, pp_start)].append(r)
            
    for (ee_code, pp_start), rows in ee_groups.items():
        # Representing exactly 1 distinct paycheck / row per employee-period
        
        # Base metadata from the first row in the group
        base_row = rows[0]
        out_row = {}
        out_row[UZIO_EMPLOYEE_ID_COL] = ee_code
        out_row[UZIO_FULL_NAME_COL] = reformat_name(str(base_row.get('NAME', '')))
        out_row[UZIO_PP_START_COL] = pp_start
        out_row[UZIO_PP_END_COL] = parse_date(base_row.get('PERIOD ENDING DATE'))
        out_row[UZIO_PAYCHECK_DATE_COL] = parse_date(base_row.get('PAY DATE'))
        
        net_pay_total = 0.0
        gross_earnings = 0.0
        total_ee_taxes = 0.0
        total_deductions = 0.0
        total_er_taxes = 0.0
        
        for row in rows:
            # We add up across all rows for this pay period (in case there are splits like CK1, CK2)
            
            # Add up Net Pay specifically
            net_amt = row.get('NET PAY', 0)
            try:
                net_amt = float(net_amt) if pd.notna(net_amt) else 0.0
            except:
                net_amt = 0.0
            net_pay_total += net_amt
            
            # Process all mapped columns
            for adp_col, target_col in mapping.items():
                amt = row.get(adp_col, 0)
                try:
                    amt = float(amt) if pd.notna(amt) else 0.0
                except:
                    amt = 0.0
                    
                if amt == 0:
                    continue
                    
                # Assign to target col
                out_row[target_col] = out_row.get(target_col, 0) + amt
                
                # Check for validation
                cat = get_adp_category(adp_col)
                if cat == 'Earnings':
                    gross_earnings += amt
                elif cat == 'Employee Taxes':
                    total_ee_taxes += amt
                elif cat == 'Employer Taxes':
                    total_er_taxes += amt
                elif cat == 'Deductions' or cat == 'Contributions':
                    total_deductions += amt
                    
            # Check skipped columns for non-zero values
            for col in row.keys():
                if col not in mapping and get_adp_category(col) != '_SKIP':
                    amt = row.get(col, 0)
                    try:
                        amt = float(amt) if pd.notna(amt) else 0.0
                    except:
                        amt = 0.0
                    if amt != 0:
                        skipped_items.append((ee_code, pp_start, col, amt))
        
        # Set final Net Pay directly into the mapped destination
        if net_pay_col_idx:
            out_row[net_pay_col_idx] = net_pay_total
            
        # Validation
        expected_net = gross_earnings - total_ee_taxes - total_deductions
        diff = round(abs(expected_net - net_pay_total), 2)
        if diff > 0.02:
            validation_results.append({
                'Employee ID': ee_code,
                'Pay Period': f"{pp_start} - {out_row[UZIO_PP_END_COL]}",
                'Pushed Gross Earnings': round(gross_earnings, 2),
                'Pushed EE Taxes': round(total_ee_taxes, 2),
                'Pushed Deductions': round(total_deductions, 2),
                'Expected Net': round(expected_net, 2),
                'Actual Source Net Pay': round(net_pay_total, 2),
                'Difference': round(expected_net - net_pay_total, 2),
            })
            
        output_rows.append(out_row)
        
    output_rows.sort(key=lambda r: (str(r.get(UZIO_EMPLOYEE_ID_COL, '')), str(r.get(UZIO_PP_START_COL, ''))))
    return output_rows, skipped_items, validation_results


def write_output_excel(uzio_wb, uzio_ws, output_rows, uzio_col_headers):
    if uzio_ws.max_row >= UZIO_DATA_START_ROW:
        uzio_ws.delete_rows(UZIO_DATA_START_ROW, uzio_ws.max_row - UZIO_DATA_START_ROW + 1)
        
    max_col = max(uzio_col_headers.keys()) if uzio_col_headers else 86
    
    for row_idx, out_row in enumerate(output_rows):
        excel_row = UZIO_DATA_START_ROW + row_idx
        for col_idx in range(1, max_col + 1):
            val = out_row.get(col_idx)
            if val is not None:
                uzio_ws.cell(row=excel_row, column=col_idx, value=val)
                
    buf = io.BytesIO()
    uzio_wb.save(buf)
    buf.seek(0)
    return buf

# ─────────────────────────────────────────────────────────────
# Main UI
# ─────────────────────────────────────────────────────────────

def render_ui():
    st.title(APP_TITLE)
    st.markdown("""
    **Convert ADP Prior Payroll History files into the Uzio Prior Payroll Template format.**
    
    1. Upload the **blank Uzio Prior Payroll Template** (headers only).
    2. Upload up to **10 ADP Prior Payroll files** (.xlsx).
    3. **Map** each dynamic ADP column to a Uzio column (or skip).
    4. **Generate** the filled template and download it.
    """)
    
    client_name = st.text_input("Client Name", value="", key="adp_pp_client_name", help="Used in output filename")
    
    st.markdown("---")
    st.markdown("### Step 1: Upload Files")
    
    col_uz, col_pc = st.columns(2)
    with col_uz:
        uzio_file = st.file_uploader("📄 Uzio Prior Payroll Template (headers only)", type=["xlsx"], key="adp_pp_uzio")
    with col_pc:
        adp_files = st.file_uploader("📁 ADP Prior Payroll History (up to 10)", type=["xlsx"], accept_multiple_files=True, key="adp_pp_files")
        
    if not uzio_file or not adp_files:
        st.info("Please upload both the Uzio template and at least one ADP file.")
        return
        
    if len(adp_files) > 10:
        st.error("Maximum 10 ADP files allowed. Please remove some files.")
        return
        
    # Read files
    try:
        uzio_bytes = io.BytesIO(uzio_file.read())
        uzio_file.seek(0)
        section_headers, uzio_col_headers, uzio_wb, uzio_ws = read_uzio_template(uzio_bytes)
    except Exception as e:
        st.error(f"Error reading Uzio template: {e}")
        return
        
    try:
        adp_rows, adp_dynamic_cols, file_summaries = read_adp_files(adp_files)
    except Exception as e:
        st.error(f"Error reading ADP files: {e}")
        return
        
    st.markdown("#### Parsed ADP Files")
    if file_summaries:
        st.dataframe(pd.DataFrame(file_summaries), hide_index=True, use_container_width=True)
    else:
        st.warning("No valid data rows found in ADP source.")
        return
        
    # Mapping UI
    st.markdown("---")
    st.markdown("### Step 2: Configure Mapping")
    st.markdown("Map each parsed ADP column to a Uzio column. Select **Skip** for items to ignore.")
    
    uzio_options = [SKIP_LABEL]
    for col_idx in sorted(uzio_col_headers.keys()):
        if col_idx >= UZIO_FIRST_DATA_COL:
            uzio_options.append(f"Col {col_idx}: {uzio_col_headers[col_idx]}")
            
    net_pay_col_idx = None
    for col_idx, hdr in uzio_col_headers.items():
        if 'net pay' in hdr.lower():
            net_pay_col_idx = col_idx
            break
            
    section_items = defaultdict(list)
    for col in sorted(adp_dynamic_cols):
        cat = get_adp_category(col)
        if cat != '_SKIP':
            section_items[cat].append(col)
            
    mapping = {}
    tabs = st.tabs(UI_SECTIONS_ORDER)
    
    for tab, section_name in zip(tabs, UI_SECTIONS_ORDER):
        with tab:
            items = section_items.get(section_name, [])
            if not items:
                st.info(f"No ADP columns found mapping to this section.")
                continue
                
            pre_filled_targets = [auto_guess_mapping(c, uzio_col_headers) for c in items]
            
            df_map = pd.DataFrame({
                'Source ADP Column': items,
                'Map To Uzio Column': pre_filled_targets,
            })
            
            edited = st.data_editor(
                df_map,
                column_config={
                    'Source ADP Column': st.column_config.Column(disabled=True, width="medium"),
                    'Map To Uzio Column': st.column_config.SelectboxColumn(
                        "Target Uzio Column", options=uzio_options, required=True, width="large"
                    ),
                },
                hide_index=True, use_container_width=True, key=f"adp_pp_map_{section_name.replace(' ', '_')}"
            )
            
            for idx, row in edited.iterrows():
                adp_col = items[idx]
                selected = row['Map To Uzio Column']
                if selected and selected != SKIP_LABEL:
                    m = re.match(r'Col (\d+):', selected)
                    if m:
                        mapping[adp_col] = int(m.group(1))
                        
    # Summarize maps
    mapped_count = len(mapping)
    total_items = sum(len(v) for v in section_items.values())
    skipped_count = total_items - mapped_count
    
    c1, c2, c3 = st.columns(3)
    c1.metric("Total ADP Items", total_items)
    c2.metric("Mapped", mapped_count)
    c3.metric("Skipped", skipped_count)
    
    if mapped_count == 0:
        st.warning("⚠️ No items mapped. Please map at least one item before generating.")
        return
        
    if net_pay_col_idx:
        st.info(f"ℹ️ Core 'NET PAY' column from ADP will automatically route to **{uzio_col_headers.get(net_pay_col_idx, 'Net Pay')}**.")
        
    with st.expander("👀 View Final Mapping Setup", expanded=False):
        review_data = []
        for col in sorted(adp_dynamic_cols):
            cat = get_adp_category(col)
            if cat == '_SKIP':
                target = f"Auto-Skipped (Category: {col})"
            else:
                target_idx = mapping.get(col)
                target = f"Col {target_idx}: {uzio_col_headers.get(target_idx, '')}" if target_idx else SKIP_LABEL
            review_data.append({'ADP Source Column': col, 'Uzio Target': target})
        st.dataframe(pd.DataFrame(review_data), hide_index=True)
        
    st.markdown("---")
    st.markdown("### Step 3: Generate & Download")
    
    if st.button("🚀 Generate Prior Payroll", type="primary", key="adp_pp_gen_btn"):
        with st.spinner("Generating..."):
            try:
                out_rows, skipped, val_results = generate_output(adp_rows, mapping, uzio_col_headers, net_pay_col_idx)
                
                s1, s2, s3 = st.columns(3)
                s1.metric("Output Rows", len(out_rows))
                s2.metric("Unique Employees", len(set(r[UZIO_EMPLOYEE_ID_COL] for r in out_rows)))
                s3.metric("Pay Periods", len(set((r[UZIO_PP_START_COL], r[UZIO_PP_END_COL]) for r in out_rows)))
                
                if val_results:
                    st.warning(f"⚠️ **{len(val_results)} employee-period(s)** have a Net Pay mismatch (Gross − Taxes − Deductions ≠ Net Pay).")
                    with st.expander(f"View Validation Summary", expanded=False):
                        df_val = pd.DataFrame(val_results)
                        st.dataframe(df_val, hide_index=True)
                    csv_b = io.BytesIO()
                    df_val.to_csv(csv_b, index=False)
                    st.download_button("📥 Download Issue Validation Report", csv_b.getvalue(), "ADP_Validation.csv", "text/csv")
                else:
                    st.success("✅ Validation passed - Net Pay aligns globally!")
                    
                if skipped:
                    st.info(f"ℹ️ **{len(skipped)} item(s)** with non-zero amounts were skipped.")
                    with st.expander(f"View Skipped Values", expanded=False):
                        st.dataframe(pd.DataFrame(skipped, columns=['ID', 'Period', 'Column', 'Amount']), hide_index=True)
                        
                uzio_file.seek(0)
                _, _, out_wb, out_ws = read_uzio_template(io.BytesIO(uzio_file.read()))
                buf = write_output_excel(out_wb, out_ws, out_rows, uzio_col_headers)
                
                ts = pd.Timestamp.now().strftime('%Y%m%d_%H%M')
                fn = f"Uzio_Prior_Payroll_{client_name}_{ts}.xlsx" if client_name else f"Uzio_Prior_Payroll_{ts}.xlsx"
                st.download_button("📥 Download Filled Uzio Template", buf.getvalue(), fn.replace(" ", "_"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                
            except Exception as e:
                st.error(f"Generation failed: {e}")
