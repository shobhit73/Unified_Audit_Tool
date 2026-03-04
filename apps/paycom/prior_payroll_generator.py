import io
import re
import pandas as pd
import streamlit as st
import openpyxl
from collections import defaultdict
from datetime import datetime

APP_TITLE = "Paycom → Uzio Prior Payroll Generator"

# ─── Uzio template constants ───
UZIO_HEADER_ROW = 5          # Row 5 holds column headers
UZIO_SECTION_ROW = 4         # Row 4 holds section headers
UZIO_DATA_START_ROW = 6      # Data starts at row 6
UZIO_EMPLOYEE_ID_COL = 1     # Col A
UZIO_FULL_NAME_COL = 2       # Col B
UZIO_SSN_COL = 3             # Col C
UZIO_PP_START_COL = 4        # Col D
UZIO_PP_END_COL = 5          # Col E
UZIO_PAYCHECK_DATE_COL = 6   # Col F
UZIO_FIRST_DATA_COL = 7      # Financial data starts at Col G

# Paycom Code Description → UI Section mapping
PAYCOM_CATEGORY_TO_SECTION = {
    'Earnings':                'Earnings',
    'W/H Taxes':               'Employee Taxes',
    'Client Side Liabilities': 'Employer Taxes',
    'Deductions':              'Deductions',      # some map to Contributions
    'Net Pay Distribution':    '_NET_PAY',         # auto-handled
    'Employee Benefits':       '_BENEFITS',        # default skip
}

UI_SECTIONS_ORDER = ['Earnings', 'Deductions', 'Contributions', 'Employee Taxes', 'Employer Taxes']

SKIP_LABEL = "── Skip (do not map) ──"


# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def parse_filename_dates(filename: str):
    """Extract pay-period start, end, and pay date from Paycom filename.
    
    Pattern examples:
      ...Pay Period 01112026 01172026Pay Date 01232026.xlsx
      ...Pay Period 01212025 01272025 Pay Date 01022026.xlsx
    """
    # Remove extension
    name = filename.rsplit('.', 1)[0]
    
    # Try pattern with optional space before "Pay Date"
    m = re.search(r'Pay Period\s+(\d{8})\s+(\d{8})\s*Pay Date\s+(\d{8})', name, re.IGNORECASE)
    if m:
        fmt = lambda d: f"{d[:2]}/{d[2:4]}/{d[4:]}"
        return fmt(m.group(1)), fmt(m.group(2)), fmt(m.group(3))
    return None, None, None


def reformat_name(raw_name: str) -> str:
    """Convert 'LAST  FIRST' → 'LAST, FIRST'."""
    if not raw_name or not isinstance(raw_name, str):
        return raw_name or ""
    parts = re.split(r'\s{2,}', raw_name.strip())
    if len(parts) == 2:
        return f"{parts[0]}, {parts[1]}"
    return raw_name.strip()


def read_uzio_template(file_bytes):
    """Read a Uzio template (headers only) and return:
      - section_headers: dict col_idx → section name (from row 4)
      - column_headers:  dict col_idx → column name (from row 5)
      - workbook:        the openpyxl workbook object
    """
    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    ws = wb[wb.sheetnames[-1]]  # last sheet is usually data
    
    # Try to detect the right sheet (look for 'Prior Payroll Data' or similar)
    for sname in wb.sheetnames:
        if 'payroll' in sname.lower() and 'instruction' not in sname.lower():
            ws = wb[sname]
            break
    
    section_headers = {}
    for cell in ws[UZIO_SECTION_ROW]:
        if cell.value:
            section_headers[cell.column] = str(cell.value).strip()
    
    column_headers = {}
    for cell in ws[UZIO_HEADER_ROW]:
        if cell.value:
            column_headers[cell.column] = str(cell.value).strip()
    
    return section_headers, column_headers, wb, ws


def read_paycom_files(uploaded_files):
    """Read all uploaded Paycom files.
    
    Returns:
      paycom_data: list of dicts with keys:
        - filename, pp_start, pp_end, pay_date, rows
        (rows is a list of dicts with keys from the header row)
      all_type_combos: set of (type_code, type_desc, code_desc)
    """
    paycom_data = []
    all_type_combos = set()
    
    for f in uploaded_files:
        pp_start, pp_end, pay_date = parse_filename_dates(f.name)
        
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        f.seek(0)
        ws = wb[wb.sheetnames[0]]
        
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            row_dict = dict(zip(headers, row))
            rows.append(row_dict)
            
            tc = row_dict.get('Type Code', '')
            td = row_dict.get('Type Description', '')
            cd = row_dict.get('Code Description', '')
            if tc:
                all_type_combos.add((str(tc).strip(), str(td).strip(), str(cd).strip()))
        
        paycom_data.append({
            'filename': f.name,
            'pp_start': pp_start,
            'pp_end': pp_end,
            'pay_date': pay_date,
            'rows': rows,
        })
    
    return paycom_data, all_type_combos


def generate_output(paycom_data, mapping, uzio_col_headers, net_pay_col_idx):
    """Generate Uzio-format output rows from Paycom data + mapping.
    
    Args:
        paycom_data: list of per-file dicts
        mapping: dict (type_code, type_desc) → uzio_col_idx (1-based) or None
        uzio_col_headers: dict col_idx → header name
        net_pay_col_idx: column index for Net Pay
    
    Returns:
        output_rows: list of dicts keyed by col_idx
        skipped_items: list of (ee_code, type_code, type_desc, amount) for skipped non-zero items
        validation_results: list of dicts with validation info per employee per pay period
    """
    output_rows = []
    skipped_items = []
    validation_results = []
    
    for pf in paycom_data:
        # Group rows by employee
        ee_groups = defaultdict(list)
        for row in pf['rows']:
            ee_code = str(row.get('EE Code', '')).strip()
            if ee_code:
                ee_groups[ee_code].append(row)
        
        for ee_code, rows in sorted(ee_groups.items()):
            # One output row per employee per pay period
            out_row = {}
            out_row[UZIO_EMPLOYEE_ID_COL] = ee_code
            out_row[UZIO_FULL_NAME_COL] = reformat_name(str(rows[0].get('EE Name', '')))
            out_row[UZIO_PP_START_COL] = pf['pp_start']
            out_row[UZIO_PP_END_COL] = pf['pp_end']
            out_row[UZIO_PAYCHECK_DATE_COL] = pf['pay_date']
            
            net_pay_total = 0.0
            gross_earnings = 0.0
            total_ee_taxes = 0.0
            total_deductions = 0.0
            total_contributions = 0.0
            total_er_taxes = 0.0
            
            for row in rows:
                tc = str(row.get('Type Code', '')).strip()
                td = str(row.get('Type Description', '')).strip()
                cd = str(row.get('Code Description', '')).strip()
                amt = row.get('Amount', 0) or 0
                try:
                    amt = float(amt)
                except (ValueError, TypeError):
                    amt = 0.0
                
                # Handle Net Pay Distribution — always sum
                if cd == 'Net Pay Distribution':
                    net_pay_total += amt
                    continue
                
                # Handle Employee Benefits — skip
                if cd == 'Employee Benefits':
                    continue
                
                # Look up mapping
                key = (tc, td)
                target_col = mapping.get(key)
                
                if target_col is None:
                    if amt != 0:
                        skipped_items.append((ee_code, pf['pp_start'], tc, td, cd, amt))
                    continue
                
                # Place amount in target column
                out_row[target_col] = out_row.get(target_col, 0) + amt
                
                # Accumulate for validation by section
                section = PAYCOM_CATEGORY_TO_SECTION.get(cd, '')
                if section == 'Earnings':
                    gross_earnings += amt
                elif section == 'Employee Taxes':
                    total_ee_taxes += amt
                elif section == 'Employer Taxes':
                    total_er_taxes += amt
                elif section == 'Deductions':
                    # Could be a deduction or contribution — check by target column
                    total_deductions += amt
            
            # Set Net Pay
            if net_pay_col_idx:
                out_row[net_pay_col_idx] = net_pay_total
            
            # Validation: Gross - EE Taxes - Deductions ≈ Net Pay
            expected_net = gross_earnings - total_ee_taxes - total_deductions
            diff = round(abs(expected_net - net_pay_total), 2)
            if diff > 0.02:  # allow 2 cent rounding tolerance
                validation_results.append({
                    'Employee ID': ee_code,
                    'Pay Period': f"{pf['pp_start']} - {pf['pp_end']}",
                    'Gross Earnings': round(gross_earnings, 2),
                    'Employee Taxes': round(total_ee_taxes, 2),
                    'Deductions': round(total_deductions, 2),
                    'Expected Net': round(expected_net, 2),
                    'Actual Net Pay': round(net_pay_total, 2),
                    'Difference': round(expected_net - net_pay_total, 2),
                })
            
            output_rows.append(out_row)
    
    # Sort by Employee ID then Pay Period Start
    output_rows.sort(key=lambda r: (r.get(UZIO_EMPLOYEE_ID_COL, ''), r.get(UZIO_PP_START_COL, '')))
    
    return output_rows, skipped_items, validation_results


def write_output_excel(uzio_wb, uzio_ws, output_rows, uzio_col_headers):
    """Write output rows into a copy of the Uzio template workbook.
    
    Preserves rows 1-5, writes data starting at row 6.
    """
    # Clear any existing data rows
    if uzio_ws.max_row >= UZIO_DATA_START_ROW:
        uzio_ws.delete_rows(UZIO_DATA_START_ROW, uzio_ws.max_row - UZIO_DATA_START_ROW + 1)
    
    max_col = max(uzio_col_headers.keys()) if uzio_col_headers else 86
    
    for row_idx, out_row in enumerate(output_rows):
        excel_row = UZIO_DATA_START_ROW + row_idx
        for col_idx in range(1, max_col + 1):
            val = out_row.get(col_idx)
            if val is not None:
                uzio_ws.cell(row=excel_row, column=col_idx, value=val)
    
    # Save to bytes
    buf = io.BytesIO()
    uzio_wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────
# Main UI
# ─────────────────────────────────────────────────────────────

def render_ui():
    st.title(APP_TITLE)
    st.markdown("""
    **Convert Paycom Prior Payroll files into the Uzio Prior Payroll Template format.**
    
    1. Upload the **blank Uzio Prior Payroll Template** (headers only).
    2. Upload up to **10 Paycom Prior Payroll files** (.xlsx).
    3. **Map** each Paycom item to a Uzio column (or skip).
    4. **Generate** the filled template and download it.
    """)
    
    client_name = st.text_input("Client Name", value="", key="pp_client_name",
                                 help="Used in the output filename")
    
    st.markdown("---")
    
    # ── STEP 1: File Upload ──────────────────────────────────
    st.markdown("### Step 1: Upload Files")
    
    col_uz, col_pc = st.columns(2)
    
    with col_uz:
        uzio_file = st.file_uploader(
            "📄 Uzio Prior Payroll Template (blank, headers only)",
            type=["xlsx"], key="pp_uzio_upload"
        )
    
    with col_pc:
        paycom_files = st.file_uploader(
            "📁 Paycom Prior Payroll Files (up to 10)",
            type=["xlsx"], accept_multiple_files=True, key="pp_paycom_upload"
        )
    
    if not uzio_file or not paycom_files:
        st.info("Please upload both the Uzio template and at least one Paycom file to continue.")
        return
    
    if len(paycom_files) > 10:
        st.error("Maximum 10 Paycom files allowed. Please remove some files.")
        return
    
    # ── Read Uzio template ──
    try:
        uzio_bytes = io.BytesIO(uzio_file.read())
        uzio_file.seek(0)
        section_headers, uzio_col_headers, uzio_wb, uzio_ws = read_uzio_template(uzio_bytes)
    except Exception as e:
        st.error(f"Error reading Uzio template: {e}")
        return
    
    # ── Read Paycom files ──
    try:
        paycom_data, all_type_combos = read_paycom_files(paycom_files)
    except Exception as e:
        st.error(f"Error reading Paycom files: {e}")
        return
    
    # ── Display parsed file summary ──
    st.markdown("#### Parsed Paycom Files")
    summary_rows = []
    for pf in sorted(paycom_data, key=lambda x: x['pp_start'] or ''):
        ee_codes = set(str(r.get('EE Code', '')).strip() for r in pf['rows'] if r.get('EE Code'))
        summary_rows.append({
            'Filename': pf['filename'],
            'Pay Period Start': pf['pp_start'],
            'Pay Period End': pf['pp_end'],
            'Pay Date': pf['pay_date'],
            'Employees': len(ee_codes),
            'Records': len(pf['rows']),
        })
    st.dataframe(pd.DataFrame(summary_rows), hide_index=True, use_container_width=True)
    
    # ── STEP 2: Mapping Configuration ──
    st.markdown("---")
    st.markdown("### Step 2: Configure Mapping")
    st.markdown("Map each Paycom item to a Uzio column. Select **Skip** for items that should not be included.")
    
    # Build dropdown options from Uzio headers (cols 7+)
    uzio_options = [SKIP_LABEL]
    for col_idx in sorted(uzio_col_headers.keys()):
        if col_idx >= UZIO_FIRST_DATA_COL:
            uzio_options.append(f"Col {col_idx}: {uzio_col_headers[col_idx]}")
    
    # Identify Net Pay column
    net_pay_col_idx = None
    for col_idx, hdr in uzio_col_headers.items():
        if 'net pay' in hdr.lower():
            net_pay_col_idx = col_idx
            break
    
    # Group type combos by UI section
    section_items = defaultdict(list)
    for tc, td, cd in sorted(all_type_combos, key=lambda x: (x[2], x[0], x[1])):
        section = PAYCOM_CATEGORY_TO_SECTION.get(cd, 'Deductions')
        if section.startswith('_'):
            continue  # Skip Net Pay Distribution and Employee Benefits
        section_items[section].append((tc, td, cd))
    
    # Build mapping via data editors in tabs
    mapping = {}   # (type_code, type_desc) → col_idx or None
    
    tabs = st.tabs(UI_SECTIONS_ORDER)
    
    for tab, section_name in zip(tabs, UI_SECTIONS_ORDER):
        with tab:
            items = section_items.get(section_name, [])
            if not items:
                st.info(f"No Paycom items found for this section.")
                continue
            
            # Build a dataframe for the data editor
            df_map = pd.DataFrame({
                'Type Code': [tc for tc, td, cd in items],
                'Description': [td for tc, td, cd in items],
                'Category': [cd for tc, td, cd in items],
                'Map To Uzio Column': [SKIP_LABEL for _ in items],
            })
            
            edited = st.data_editor(
                df_map,
                column_config={
                    'Type Code': st.column_config.Column(disabled=True, width="small"),
                    'Description': st.column_config.Column(disabled=True, width="medium"),
                    'Category': st.column_config.Column(disabled=True, width="medium"),
                    'Map To Uzio Column': st.column_config.SelectboxColumn(
                        "Target Uzio Column",
                        options=uzio_options,
                        required=True,
                        width="large"
                    ),
                },
                hide_index=True,
                use_container_width=True,
                key=f"pp_map_{section_name.replace(' ', '_').lower()}"
            )
            
            # Extract mapping from edited dataframe
            for idx, row in edited.iterrows():
                tc = items[idx][0]
                td = items[idx][1]
                selected = row['Map To Uzio Column']
                if selected and selected != SKIP_LABEL:
                    # Parse "Col XX: Header Name" → col_idx
                    col_match = re.match(r'Col (\d+):', selected)
                    if col_match:
                        mapping[(tc, td)] = int(col_match.group(1))
    
    # Show mapping summary
    mapped_count = len(mapping)
    total_items = sum(len(v) for v in section_items.values())
    skipped_count = total_items - mapped_count
    
    col_m1, col_m2, col_m3 = st.columns(3)
    col_m1.metric("Total Paycom Items", total_items)
    col_m2.metric("Mapped", mapped_count)
    col_m3.metric("Skipped", skipped_count)
    
    if mapped_count == 0:
        st.warning("⚠️ No items are mapped. Please configure mappings above before generating.")
        return
    
    if net_pay_col_idx:
        st.info(f"ℹ️ Net Pay Distribution items (NET, NP1, NP2, etc.) will be automatically **summed** into column **{uzio_col_headers.get(net_pay_col_idx, 'Net Pay')}**.")
    else:
        st.warning("⚠️ Could not auto-detect a 'Net Pay' column in the Uzio template. Net pay will not be populated.")
    
    # ── STEP 3: Generate ────────────────────────────────────
    st.markdown("---")
    st.markdown("### Step 3: Generate & Download")
    
    if st.button("🚀 Generate Prior Payroll", type="primary", key="pp_generate_btn"):
        with st.spinner("Generating prior payroll..."):
            try:
                output_rows, skipped_items, validation_results = generate_output(
                    paycom_data, mapping, uzio_col_headers, net_pay_col_idx
                )
                
                # ── Statistics ──
                unique_ees = set(r.get(UZIO_EMPLOYEE_ID_COL) for r in output_rows)
                unique_periods = set((r.get(UZIO_PP_START_COL), r.get(UZIO_PP_END_COL)) for r in output_rows)
                
                scol1, scol2, scol3 = st.columns(3)
                scol1.metric("Output Rows", len(output_rows))
                scol2.metric("Unique Employees", len(unique_ees))
                scol3.metric("Pay Periods", len(unique_periods))
                
                # ── Validation results ──
                if validation_results:
                    st.warning(f"⚠️ **{len(validation_results)} employee-pay period(s)** have a Net Pay mismatch (Gross − Taxes − Deductions ≠ Net Pay).")
                    with st.expander(f"View {len(validation_results)} Validation Mismatch(es)", expanded=False):
                        df_val = pd.DataFrame(validation_results)
                        st.dataframe(df_val, hide_index=True, use_container_width=True)
                    
                    val_csv = io.BytesIO()
                    df_val.to_csv(val_csv, index=False)
                    val_csv.seek(0)
                    st.download_button(
                        "📥 Download Validation Report (CSV)",
                        data=val_csv.getvalue(),
                        file_name=f"Validation_Mismatches_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.csv",
                        mime="text/csv",
                        key="pp_val_dl"
                    )
                else:
                    st.success("✅ Validation passed — all Net Pay totals match!")
                
                # ── Skipped items summary ──
                if skipped_items:
                    st.info(f"ℹ️ **{len(skipped_items)} line item(s)** with non-zero amounts were **skipped** (unmapped).")
                    with st.expander(f"View {len(skipped_items)} Skipped Item(s)", expanded=False):
                        df_skip = pd.DataFrame(skipped_items, columns=[
                            'Employee ID', 'Pay Period Start', 'Type Code', 'Description', 'Category', 'Amount'
                        ])
                        st.dataframe(df_skip, hide_index=True, use_container_width=True)
                
                # ── Write output Excel ──
                # Re-read the template to get a clean copy
                uzio_file.seek(0)
                uzio_bytes2 = io.BytesIO(uzio_file.read())
                _, _, out_wb, out_ws = read_uzio_template(uzio_bytes2)
                
                output_buf = write_output_excel(out_wb, out_ws, output_rows, uzio_col_headers)
                
                timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M')
                fname = f"Uzio_Prior_Payroll_{client_name}_{timestamp}.xlsx" if client_name else f"Uzio_Prior_Payroll_{timestamp}.xlsx"
                fname = fname.replace(' ', '_')
                
                st.download_button(
                    "📥 Download Filled Uzio Template",
                    data=output_buf.getvalue(),
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="pp_output_dl"
                )
                
                st.success("✅ Prior Payroll generated successfully!")
                
            except Exception as e:
                st.error(f"Error generating output: {e}")
                import traceback
                st.code(traceback.format_exc())
