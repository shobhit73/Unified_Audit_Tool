import pandas as pd
import streamlit as st
import io
import re
import numpy as np
from datetime import datetime, date

# --- Hardcoded Mappings ---

# Map Uzio Raw Headers -> Internal Standard Names
UZIO_RAW_MAPPING = {
    'Employee ID*': 'Employee ID',
    'Employee First Name*': 'First Name',
    'Employee Last Name*': 'Last Name',
    'Employee Middle Initial': 'Middle Initial',
    'Employee Suffix': 'Suffix',
    'Employment Status*': 'Employment Status',
    'Date of Hire*': 'Hire Date',
    'Original DOH': 'Original Hire Date',
    'Termination Date': 'Termination Date',
    'Termination Reason': 'Termination Reason',
    'Employment Type*': 'Employment Type',
    'Pay Type*': 'Pay Type',
    'Annual Salary(Digits)**': 'Annual Salary',
    'Hourly Pay Rate**': 'Hourly Pay Rate',
    'Working Hours per Week(Digits)**': 'Working Hours',
    'Job Title': 'Job Title',
    'Department': 'Department',
    'Official Email*': 'Work Email',
    'Personal Email': 'Personal Email',
    'Phone Number(Digits)': 'Phone Number',
    'Employee SSN': 'SSN',
    'Employee Date of Birth*': 'DOB',
    'Employee Gender*': 'Gender',
    'Employee Tobacco usage in last 12 months': 'Tobacco User',
    'FLSA Classification': 'FLSA Classification',
    'Employee Address Line 1': 'Address Line 1',
    'Employee Address Line 2': 'Address Line 2',
    'City*': 'City',
    'Zipcode*': 'Zip',
    'State(Abbreviation)*': 'State',
    'Mailing Address Line 1': 'Mailing Address Line 1',
    'Mailing Address Line 2': 'Mailing Address Line 2',
    'Mailing City': 'Mailing City',
    'Mailing Zipcode': 'Mailing Zip',
    'Mailing State(Abbreviation)': 'Mailing State',
    'Reporting Manager ID': 'Reports To ID',
    'Work Location': 'Work Location',
    'License Number*': 'License Number',
    'License Expiration Date': 'License Expiration Date'
}

def read_uzio_raw_file(uploaded_file):
    """
    Reads the raw Uzio .xlsm export.
    Expects 'Employee Details' sheet.
    Headers are in Row 4 (Index 3).
    Renames columns to Internal Standard Names.
    """
    try:
        # Read Excel - header=3 means 4th row is header
        df = pd.read_excel(uploaded_file, sheet_name='Employee Details', header=3)
        
        # Strip whitespace and replace newlines/multiple spaces with single space
        df.columns = df.columns.astype(str).str.replace(r'\s+', ' ', regex=True).str.strip()
        
        # Rename columns based on mapping
        # Only rename columns that exist in the mapping
        rename_dict = {k: v for k, v in UZIO_RAW_MAPPING.items() if k in df.columns}
        df = df.rename(columns=rename_dict)
        
        # Ensure 'Employee ID' is string (remove decimals if any)
        if 'Employee ID' in df.columns:
            df['Employee ID'] = df['Employee ID'].astype(str).str.replace(r'\.0$', '', regex=True)
            
        print("Uzio Raw File Read Successfully.")
        print(f"Columns Found: {list(df.columns)}")
        return df

    except Exception as e:
        st.error(f"Error reading Uzio Raw File: {e}")
        return None

def extract_mappings_from_uzio(df_source, df_template, vendor_map):
    """
    Extracts Job Title and Work Location mappings from a pre-filled Uzio template.
    Matches employees between source and template using Employee ID.
    Returns: (job_mappings, loc_mappings) dictionaries.
    """
    job_mappings = {}
    loc_mappings = {}
    
    src_id_col = vendor_map.get('Employee ID')
    src_job_col = vendor_map.get('Job Title')
    src_loc_col = vendor_map.get('Work Location')
    
    if not src_id_col or src_id_col not in df_source.columns:
        return job_mappings, loc_mappings
        
    # Uzio headers (Standard names after read_uzio_raw_file)
    # Note: UZIO_RAW_MAPPING maps 'Employee ID*' to 'Employee ID'
    UZIO_ID = 'Employee ID'
    UZIO_JOB = 'Job Title'
    UZIO_LOC = 'Work Location'
    
    if UZIO_ID not in df_template.columns or UZIO_JOB not in df_template.columns or UZIO_LOC not in df_template.columns:
        # Try raw names just in case it wasn't processed
        if 'Employee ID*' in df_template.columns: UZIO_ID = 'Employee ID*'
        else: return job_mappings, loc_mappings
        
    # Build a lookup for Uzio records
    uzio_lookup = {}
    for _, t_row in df_template.iterrows():
        tid = str(t_row.get(UZIO_ID, "")).strip().replace(".0", "")
        if tid:
            uzio_lookup[tid] = {
                'job': str(t_row.get(UZIO_JOB, "")).strip(),
                'loc': str(t_row.get(UZIO_LOC, "")).strip()
            }
            
    # Iterate source and build mapping
    for _, s_row in df_source.iterrows():
        sid = str(s_row.get(src_id_col, "")).strip().replace(".0", "")
        if not sid or sid not in uzio_lookup:
            continue
            
        u_job = uzio_lookup[sid]['job']
        u_loc = uzio_lookup[sid]['loc']
        
        s_job = str(s_row.get(src_job_col, "")).strip() if src_job_col else ""
        s_loc = str(s_row.get(src_loc_col, "")).strip() if src_loc_col else ""
        
        if s_job and u_job and u_job.lower() not in ['nan', 'none', '']:
            job_mappings[s_job] = u_job
        if s_loc and u_loc and u_loc.lower() not in ['nan', 'none', '']:
            loc_mappings[s_loc] = u_loc
                
    return job_mappings, loc_mappings

def norm_col(c):
    """Normalize column names to be case-insensitive and stripped."""
    if c is None: return ""
    return str(c).strip().replace("\n", " ").strip()

def norm_colname(c: str) -> str:
    """Robust column normalization: handles newlines, special quotes, and bracketed suffixes."""
    if c is None:
        return ""
    c = str(c).replace("\n", " ").replace("\r", " ")
    c = c.replace("\u00A0", " ")
    c = c.replace("’", "'").replace("“", '"').replace("”", '"')
    # Remove bracketed suffixes like (Personal Profile) or (Employment Profile - Pay Rates)
    c = re.sub(r'\(.*?\)', '', c)
    c = re.sub(r"\s+", " ", c).strip()
    c = c.replace("*", "")
    c = c.strip('"').strip("'")
    return c

def norm_key_series(s: pd.Series) -> pd.Series:
    """Normalize a pandas Series of keys (like Employee IDs), handling floats like '123.0' and NaN."""
    s2 = s.astype(object).where(~s.isna(), "")
    def _fix(v):
        v = str(v).strip()
        v = v.replace("\u00A0", " ")
        if re.fullmatch(r"\d+\.0+", v):
            v = v.split(".")[0]
        return v
    return s2.map(_fix)

def norm_blank(x):
    """Normalize blank/NaN values to an empty string."""
    if x is None:
        return ""
    if isinstance(x, float) and np.isnan(x):
        return ""
    if isinstance(x, str) and x.strip().lower() in {"", "nan", "none", "null"}:
        return ""
    return x

def try_parse_date(x):
    """Attempt to parse various date formats into standard MM/DD/YYYY string."""
    x = norm_blank(x)
    if x == "":
        return ""
    if isinstance(x, (datetime, date, np.datetime64, pd.Timestamp)):
        ts = pd.to_datetime(x, errors='coerce')
        if pd.isna(ts):
            return ""
        return ts.strftime("%m/%d/%Y")
    if isinstance(x, str):
        s = x.strip()
        try:
            ts = pd.to_datetime(s, errors='coerce')
            if pd.isna(ts):
                return s  # return the original string if it can't be parsed
            return ts.strftime("%m/%d/%Y")
        except Exception:
            return s
    return str(x)

def ensure_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    If multiple columns normalize to the same name (case-insensitive), keep only the first one.
    This prevents .loc[index, col] from returning a Series instead of a scalar.
    """
    df = df.copy()
    norm_cols = [norm_colname(c).casefold() for c in df.columns]
    seen = set()
    to_keep = []
    for i, nc in enumerate(norm_cols):
        if nc not in seen:
            seen.add(nc)
            to_keep.append(i)
    return df.iloc[:, to_keep]

def safe_val(df, idx, col):
    """Safely get a scalar value from a dataframe, even if duplicate columns somehow exist."""
    if idx is None or col not in df.columns:
        return ""
    val = df.loc[idx, col]
    if isinstance(val, pd.Series):
        return val.iloc[0]
    return val

def normalize_space_and_case(x):
    """Strip whitespace, collapse consecutive spaces, and convert to casefold."""
    x = norm_blank(x)
    if x == "":
        return ""
    s = str(x).strip()
    s = s.replace("\u00A0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s.casefold()

def find_col(df_cols, *candidate_names):
    """Find a column in a list of columns by matching against several candidates (case-insensitive)."""
    norm_map = {norm_colname(c).casefold(): c for c in df_cols}
    for cand in candidate_names:
        key = norm_colname(cand).casefold()
        if key in norm_map:
            return norm_map[key]
    return None

def as_float_or_none(x):
    """Convert a value to float or return None if not a number."""
    x = norm_blank(x)
    if x == "":
        return None
    if isinstance(x, (int, float, np.integer, np.floating)):
        try:
            return float(x)
        except Exception:
            return None
    if isinstance(x, str):
        s = x.strip().replace(",", "").replace("$", "")
        if s == "":
            return None
        try:
            return float(s)
        except Exception:
            return None
    return None

# --- Job Titles that Uzio ALWAYS treats as Hourly/Non-Exempt ---
# Any salaried employee with one of these job titles must be flagged.
HOURLY_ONLY_JOB_TITLES = {
    "driver",
    "lead driver",
    "walker",
    "helper",
    "driver-lite",
    "driver-step van",
    "driver-unscheduled",
    "ddu dedicated",
    "ddu shared",
    "delivery associate",
    "delivery associates",
    "driver -major appliance",
}

def is_hourly_only_job_title(jt_val: str) -> bool:
    """Return True if this job title is always Hourly in Uzio."""
    jt = jt_val.strip().lower()
    if jt in HOURLY_ONLY_JOB_TITLES:
        return True
    # Also catch any title ending with 'driver' (e.g. 'amazon driver', 'flex driver')
    if jt.endswith("driver"):
        return True
    return False

def clean_money_val(x):
    """Parse money/percentage strings to float. Returns original string if not a number."""
    if pd.isna(x) or x == "":
        return 0.0
    s = str(x).strip()
    s_clean = s.replace("$", "").replace("%", "").replace(",", "")
    s_clean = s_clean.replace("(", "-").replace(")", "") # Handle accounting negative
    try:
        return float(s_clean)
    except:
        return 0.0

def validate_source_data(df_source, resolved_field_map):
    """
    Comprehensive pre-generation sanity checks on the raw source data.
    Returns a dict:
      - 'hard_errors': pd.DataFrame of blocking issues (blank mandatory fields, bad zip, missing salary)
      - 'flsa_corrections': pd.DataFrame of FLSA auto-corrections made
      - 'email_fallbacks': pd.DataFrame of email fallback auto-fills made
    The caller should block generation if hard_errors is not empty.
    """
    hard_errors = []
    flsa_corrections = []
    flsa_blanks = []
    type_blanks = []
    intern_corrections = []
    email_fallbacks = []
    salaried_drivers = []
    anomalies = [] # New: For Hourly Exempt / Salaried Non-Exempt flags
    
    # Resolve column references
    emp_id_col = resolved_field_map.get('Employee ID')
    status_col = resolved_field_map.get('Employment Status')
    type_col = resolved_field_map.get('Employment Type')
    pay_type_col = resolved_field_map.get('Pay Type')
    job_title_col = resolved_field_map.get('Job Title')
    location_col = resolved_field_map.get('Work Location')
    zip_col = resolved_field_map.get('Zip')
    salary_col = resolved_field_map.get('Annual Salary')
    flsa_col = resolved_field_map.get('FLSA Classification')
    work_email_col = resolved_field_map.get('Work Email')
    personal_email_col = resolved_field_map.get('Personal Email')
    hours_col = resolved_field_map.get('Working Hours')
    state_col = resolved_field_map.get('State')
    hire_date_col = resolved_field_map.get('Hire Date')
    term_date_col = resolved_field_map.get('Termination Date')
    first_name_col = resolved_field_map.get('First Name')
    last_name_col = resolved_field_map.get('Last Name')
    
    def get_emp_ref(row, idx):
        ref = f"Row {idx+2}"
        if emp_id_col and emp_id_col in df_source.columns:
            eid = row.get(emp_id_col)
            if pd.notna(eid) and str(eid).strip():
                ref = str(eid).strip()
        return ref

    def get_emp_name(row):
        fname = ""
        lname = ""
        if first_name_col and first_name_col in df_source.columns:
            val = row.get(first_name_col)
            fname = str(val).strip() if pd.notna(val) else ""
        if last_name_col and last_name_col in df_source.columns:
            val = row.get(last_name_col)
            lname = str(val).strip() if pd.notna(val) else ""
        return f"{fname} {lname}".strip()
    
    for idx, row in df_source.iterrows():
        emp_ref = get_emp_ref(row, idx)
        
        # --- HARD STOP CHECKS ---
        missing = []
        
        # 0. Blank SSN
        ssn_col = resolved_field_map.get('SSN')
        if ssn_col and ssn_col in df_source.columns:
            ssn_val = row.get(ssn_col)
            if pd.isna(ssn_val) or str(ssn_val).strip() == "":
                missing.append("SSN (blank)")

        # 1. Blank Employment Status
        if status_col and status_col in df_source.columns:
            val = row.get(status_col)
            if pd.isna(val) or str(val).strip() == "":
                missing.append("Employment Status")
        
        # 2. Blank Employment Type -> Added to hard errors as per user request
        if type_col and type_col in df_source.columns:
            val = row.get(type_col)
            if pd.isna(val) or str(val).strip() == "":
                missing.append("Employment Type")
                type_blanks.append({
                    'Employee ID': emp_ref,
                    'Original Employment Type': '(blank)'
                })
        
        # 3. Blank Pay Type
        pay_val = ""
        if pay_type_col and pay_type_col in df_source.columns:
            pay_val = row.get(pay_type_col)
            if pd.isna(pay_val) or str(pay_val).strip() == "":
                missing.append("Pay Type")
                pay_val = ""
            else:
                pay_val = str(pay_val).strip().lower()
        
        # 4. Blank Job Title
        if job_title_col and job_title_col in df_source.columns:
            val = row.get(job_title_col)
            if pd.isna(val) or str(val).strip() == "":
                missing.append("Job Title")
                
        # 4b. Blank Work Location
        if location_col and location_col in df_source.columns:
            val = row.get(location_col)
            if pd.isna(val) or str(val).strip() == "":
                missing.append("Work Location")
        
        # 5. Invalid Zip Code (must be 5 digits)
        if zip_col and zip_col in df_source.columns:
            zip_val = row.get(zip_col)
            if pd.isna(zip_val) or str(zip_val).strip() == "":
                missing.append("Zip Code (blank)")
            else:
                # Strip to digits only
                import re
                digits_only = re.sub(r'[^0-9]', '', str(zip_val).split('.')[0].split('-')[0])
                if len(digits_only) < 5:
                    missing.append(f"Zip Code ('{zip_val}' is not 5 digits)")
        
        # 6. Salaried without Annual Salary
        if pay_val and ("salary" in pay_val or "salaried" in pay_val):
            if salary_col and salary_col in df_source.columns:
                sal_val = row.get(salary_col)
                if pd.isna(sal_val) or str(sal_val).strip() == "" or str(sal_val).strip() == "0":
                    missing.append("Annual Salary (required for Salaried)")
        
        # 6b. Salaried Hourly-Only Job Title Exception — always check, irrespective of salary value
        if pay_val and ("salary" in pay_val or "salaried" in pay_val):
            if job_title_col and job_title_col in df_source.columns:
                jt_raw = row.get(job_title_col)
                jt_val = str(jt_raw).strip().lower() if pd.notna(jt_raw) else ""
                if jt_val and jt_val != "nan":
                    if is_hourly_only_job_title(jt_val):
                        missing.append(f"Salaried Hourly-Only Exception (Job Title '{str(jt_raw).strip()}' must be Hourly/Non-Exempt in Uzio)")
                        salaried_drivers.append({
                            'Employee ID': emp_ref,
                            'Job Title': str(jt_raw).strip(),
                            'Pay Type': str(row.get(pay_type_col, '')).strip()
                        })
        
        # 7. Blank Working Hours
        if hours_col and hours_col in df_source.columns:
            hrs_val = row.get(hours_col)
            if pd.isna(hrs_val) or str(hrs_val).strip() == "" or str(hrs_val).strip().lower() == 'nan':
                missing.append("Working Hours (blank)")
        
        # 8. State must be 2-character abbreviation (not full name)
        if state_col and state_col in df_source.columns:
            state_val = row.get(state_col)
            if pd.notna(state_val) and str(state_val).strip():
                sv = str(state_val).strip()
                if len(sv) > 2:
                    missing.append(f"State ('{sv}' is full name, need 2-char abbreviation)")
                    
        # 9. Termination Date vs Hire Date validity (only if Terminated/Inactive)
        if hire_date_col and hire_date_col in df_source.columns and term_date_col and term_date_col in df_source.columns:
            hire_val = row.get(hire_date_col)
            term_val = row.get(term_date_col)
            
            # Check if Employee Status is actually terminated or inactive
            is_terminated = False
            if status_col and status_col in df_source.columns:
                emp_status_val = str(row.get(status_col)).strip().lower()
                if "term" in emp_status_val or "inactive" in emp_status_val:
                    is_terminated = True
            
            if is_terminated and pd.notna(hire_val) and str(hire_val).strip() != "" and pd.notna(term_val) and str(term_val).strip() != "":
                # Attempt to parse both dates
                try:
                    # Using pd.to_datetime with errors='coerce' to safely parse strings to datetimes
                    parsed_hire = pd.to_datetime(hire_val, errors='coerce')
                    parsed_term = pd.to_datetime(term_val, errors='coerce')
                    
                    if pd.notna(parsed_hire) and pd.notna(parsed_term):
                        if parsed_term < parsed_hire:
                            missing.append(f"Date of termination ({parsed_term.strftime('%Y-%m-%d')}) predates date of hire ({parsed_hire.strftime('%Y-%m-%d')})")
                except Exception:
                    pass # Ignore if dates are malformed, we just won't flag this specific error
        
        if missing:
            hard_errors.append({
                'Employee ID': emp_ref,
                'Name': get_emp_name(row),
                'Issue': ", ".join(missing)
            })
        
        # --- SOFT CHECKS (auto-correct) ---
        
        # 7. FLSA Mismatch
        if flsa_col and flsa_col in df_source.columns and pay_type_col and pay_type_col in df_source.columns:
            flsa_val = row.get(flsa_col)
            if pd.notna(flsa_val) and str(flsa_val).strip():
                flsa_str = str(flsa_val).strip().lower()
                
                if pay_val and ("hourly" in pay_val or "hour" in pay_val):
                    if "exempt" in flsa_str and "non" not in flsa_str:
                        # Hourly should be Non-Exempt
                        flsa_corrections.append({
                            'Employee ID': emp_ref,
                            'Name': get_emp_name(row),
                            'Pay Type': str(row.get(pay_type_col, '')).strip(),
                            'Original FLSA': str(flsa_val).strip(),
                            'Corrected FLSA': 'Non-Exempt'
                        })
                        anomalies.append({
                            'Employee ID': emp_ref,
                            'Name': get_emp_name(row),
                            'Issue': f"Hourly Exempt (Pay Type: {str(row.get(pay_type_col, '')).strip()}, FLSA: {str(flsa_val).strip()})"
                        })
                elif pay_val and ("salary" in pay_val or "salaried" in pay_val):
                    if "non" in flsa_str and "exempt" in flsa_str:
                        # Salaried should be Exempt
                        flsa_corrections.append({
                            'Employee ID': emp_ref,
                            'Name': get_emp_name(row),
                            'Pay Type': str(row.get(pay_type_col, '')).strip(),
                            'Original FLSA': str(flsa_val).strip(),
                            'Corrected FLSA': 'Exempt'
                        })
                        anomalies.append({
                            'Employee ID': emp_ref,
                            'Name': get_emp_name(row),
                            'Issue': f"Salaried Non-Exempt (Pay Type: {str(row.get(pay_type_col, '')).strip()}, FLSA: {str(flsa_val).strip()})"
                        })
            else:
                # FLSA is blank — soft flag
                if pay_val and ("hourly" in pay_val or "hour" in pay_val or "salary" in pay_val or "salaried" in pay_val):
                    flsa_blanks.append({
                        'Employee ID': emp_ref,
                        'Name': get_emp_name(row),
                        'Pay Type': str(row.get(pay_type_col, '')).strip(),
                        'FLSA Classification': '(blank)'
                    })
        
        # 8. Blank Work Email → fill with Personal Email
        if work_email_col and work_email_col in df_source.columns:
            we_val = row.get(work_email_col)
            if pd.isna(we_val) or str(we_val).strip() == "":
                if personal_email_col and personal_email_col in df_source.columns:
                    pe_val = row.get(personal_email_col)
                    if pd.notna(pe_val) and str(pe_val).strip():
                        email_fallbacks.append({
                            'Employee ID': emp_ref,
                            'Name': get_emp_name(row),
                            'Personal Email Used': str(pe_val).strip()
                        })
        
        # 9. Intern in Worker Category → auto-correct to Part Time
        if type_col and type_col in df_source.columns:
            type_val = row.get(type_col)
            if pd.notna(type_val) and 'intern' in str(type_val).strip().lower():
                intern_corrections.append({
                    'Employee ID': emp_ref,
                    'Name': get_emp_name(row),
                    'Original Employment Type': str(type_val).strip(),
                    'Corrected Employment Type': 'Part Time'
                })
    
    return {
        'hard_errors': pd.DataFrame(hard_errors),
        'flsa_corrections': pd.DataFrame(flsa_corrections),
        'flsa_blanks': pd.DataFrame(flsa_blanks),
        'type_blanks': pd.DataFrame(type_blanks),
        'intern_corrections': pd.DataFrame(intern_corrections),
        'email_fallbacks': pd.DataFrame(email_fallbacks),
        'salaried_drivers': pd.DataFrame(salaried_drivers).assign(Name=lambda d: d['Employee ID'].map(lambda x: next((e['Name'] for e in hard_errors if e['Employee ID'] == x), "")) if not d.empty else ""),
        'anomalies': pd.DataFrame(anomalies)
    }

def generate_uzio_template(df_source, vendor_field_map, fix_options=None):
    """
    Generate an Uzio Census Template DataFrame from a source DataFrame.
    """
    
    # Create an empty dataframe with Uzio headers
    uzio_headers = list(UZIO_RAW_MAPPING.keys())
    df_uzio = pd.DataFrame(columns=uzio_headers)
    
    # Iterate through each Uzio expected header
    for uzio_header, std_name in UZIO_RAW_MAPPING.items():
        # Special Case: Leave blank
        if std_name in ['Job Title', 'Department', 'Work Location']:
            df_uzio[uzio_header] = ""
            continue
            
        vendor_col = vendor_field_map.get(std_name)
        if vendor_col and vendor_col in df_source.columns:
            # We have a direct mapping
            series = df_source[vendor_col].copy()
            
            # Apply formatting rules
            if std_name == 'Middle Initial':
                series = series.apply(lambda x: str(x).strip()[0] if pd.notna(x) and str(x).strip() else "")
            elif std_name in ['Hire Date', 'Original Hire Date', 'Termination Date', 'DOB']:
                def format_date(d):
                    if pd.isna(d) or str(d).strip() == "": return ""
                    try:
                        dt = pd.to_datetime(str(d).strip(), errors='coerce')
                        if pd.isna(dt): return str(d).strip()
                        return dt.strftime('%d/%m/%Y')
                    except:
                        return str(d).strip()
                series = series.apply(format_date)
            elif std_name == 'License Expiration Date':
                def format_license_exp_date(d):
                    if pd.isna(d) or str(d).strip() == "": return ""
                    d_str = str(d).strip()
                    # Never allow placeholder invalid dates
                    if '00/00/0000' in d_str or d_str in ('0', '00', '0000'): return ""
                    try:
                        dt = pd.to_datetime(d_str, errors='coerce')
                        if pd.isna(dt): return ""
                        return dt.strftime('%m/%d/%Y')  # Uzio wants MM/DD/YYYY
                    except:
                        return ""
                series = series.apply(format_license_exp_date)
            elif std_name == 'SSN':
                series = series.apply(lambda x: str(x).replace("-", "").strip() if pd.notna(x) else "")
            elif std_name == 'Gender':
                def format_gender(g):
                    if pd.isna(g) or str(g).strip() == "": return ""
                    g_str = str(g).strip().lower()
                    if g_str.startswith('m'): return "Male"
                    if g_str.startswith('f'): return "Female"
                    return ""
                series = series.apply(format_gender)
            elif std_name == 'Employment Status':
                def format_status(x):
                    if pd.isna(x): return ""
                    s = str(x).strip().lower()
                    if not s: return ""
                    
                    if fix_options and fix_options.get('fix_status', False):
                        if 'not hired' in s: return 'EXCLUDE'
                        if 'inactive' in s: return 'TERMINATED'
                        if 'leave' in s: return 'ACTIVE'
                        if 'term' in s: return 'TERMINATED'
                        if 'active' in s: return 'ACTIVE'
                    
                    return str(x).strip().upper()
                series = series.apply(format_status)
            elif std_name in ['Zip', 'Mailing Zip']:
                def format_zip(z):
                    if pd.isna(z) or str(z).strip() == "": return ""
                    # Keep digits only
                    import re
                    z_clean = re.sub(r'\D', '', str(z).strip())
                    if not z_clean: return ""
                    # Pad to 5 or truncate to 5
                    if len(z_clean) < 5:
                        return z_clean.zfill(5)
                    else:
                        return z_clean[:5]
                series = series.apply(format_zip)
            elif std_name == 'Employment Type':
                def format_emp_type(et):
                    if pd.isna(et) or str(et).strip() == "": return ""
                    et_str = str(et).strip().lower()
                    
                    if fix_options and fix_options.get('fix_type', False):
                        if 'full' in et_str: return 'Full Time'
                        if 'part' in et_str: return 'Part Time'
                        if 'season' in et_str: return 'Seasonal'
                        if 'other' in et_str: return 'Other'
                        if 'intern' in et_str: return 'Part Time'
                    
                    return str(et).strip()
                series = series.apply(format_emp_type)
            elif std_name == 'Termination Reason':
                def format_term_reason(tr):
                    if pd.isna(tr) or str(tr).strip() == "": return ""
                    tr_str = str(tr).strip().lower()
                    
                    if "involuntary" in tr_str or "invluntary" in tr_str:
                        return "Involuntary Termination of Employment"
                    if "voluntary" in tr_str or "quit" in tr_str:
                        return "Voluntary Termination of Employment"
                    if "death" in tr_str:
                        return "Death"
                    if "retire" in tr_str:
                        return "Retirement"
                    if "disability" in tr_str:
                        return "Permanent Disability"
                    if "transfer" in tr_str:
                        return "Transfer"
                    
                    # Anything else that is not blank gets 'Other'
                    return "Other"
                series = series.apply(format_term_reason)
            # We port the data
            df_uzio[uzio_header] = series
        else:
            df_uzio[uzio_header] = ""
    # Filter out excluded employees (e.g., 'not hired')
    if 'Employment Status*' in df_uzio.columns:
        df_uzio = df_uzio[df_uzio['Employment Status*'] != 'EXCLUDE'].copy()
        
    # Apply Work Email Fallback (Optional)
    if fix_options and fix_options.get('fix_emails', False):
        if 'Official Email*' in df_uzio.columns and 'Personal Email' in df_uzio.columns:
            # Fill missing Work Emails with Personal Email
            missing_work_mask = df_uzio['Official Email*'].isna() | (df_uzio['Official Email*'].astype(str).str.strip() == "")
            df_uzio.loc[missing_work_mask, 'Official Email*'] = df_uzio.loc[missing_work_mask, 'Personal Email']

    # --- License Rules (Optional) ---
    if fix_options and fix_options.get('fix_license', False):
        # Rule 1: Never allow License Expiration Date if License Number is blank
        # Rule 2: Never allow 00/00/0000 in License Expiration Date
        lic_num_col = 'License Number*'
        lic_exp_col = 'License Expiration Date'
        if lic_exp_col in df_uzio.columns:
            # Clear 00/00/0000 or similar invalid placeholders
            bad_exp_mask = df_uzio[lic_exp_col].astype(str).str.strip().isin(['00/00/0000', '0', '00', '0000', 'nan', 'NaT', ''])
            df_uzio.loc[bad_exp_mask, lic_exp_col] = ""
            # Clear expiration date if no license number
            if lic_num_col in df_uzio.columns:
                no_license_mask = df_uzio[lic_num_col].isna() | (df_uzio[lic_num_col].astype(str).str.strip() == "") | (df_uzio[lic_num_col].astype(str).str.strip() == 'nan')
                df_uzio.loc[no_license_mask, lic_exp_col] = ""

    # Apply Pay Type rules
    if 'Pay Type*' in df_uzio.columns:
        pay_type_series = df_uzio['Pay Type*'].astype(str).str.lower().str.strip()
        
        # Hourly logic
        hourly_mask = pay_type_series.str.contains('hour', na=False)
        df_uzio.loc[hourly_mask, 'Pay Type*'] = "Hourly"
        if 'Annual Salary(Digits)**' in df_uzio.columns:
            df_uzio.loc[hourly_mask, 'Annual Salary(Digits)**'] = ""
            # Apply Pay Type rules (Optional)
            if fix_options and fix_options.get('fix_flsa', False):
                # Enforce Hourly = Non-Exempt
                if 'FLSA Classification' in df_uzio.columns:
                    df_uzio.loc[hourly_mask, 'FLSA Classification'] = "Non-Exempt"
            
        # Salaried logic
        salary_mask = pay_type_series.str.contains('salar', na=False)
        df_uzio.loc[salary_mask, 'Pay Type*'] = "Salaried"
        if 'Hourly Pay Rate**' in df_uzio.columns:
            df_uzio.loc[salary_mask, 'Hourly Pay Rate**'] = 0
        if 'Working Hours per Week(Digits)**' in df_uzio.columns:
            df_uzio.loc[salary_mask, 'Working Hours per Week(Digits)**'] = ""

        # Enforce Salaried = Exempt (Optional)
        if fix_options and fix_options.get('fix_flsa', False):
            if 'FLSA Classification' in df_uzio.columns:
                df_uzio.loc[salary_mask, 'FLSA Classification'] = "Exempt"
            
        # Mandatory fallback: if FLSA is still blank, default to Non-Exempt as a safety measure (Optional)
        if fix_options and fix_options.get('fix_flsa', False):
            if 'FLSA Classification' in df_uzio.columns:
                blank_flsa_mask = df_uzio['FLSA Classification'].isna() | (df_uzio['FLSA Classification'].astype(str).str.strip() == "")
                df_uzio.loc[blank_flsa_mask, 'FLSA Classification'] = "Non-Exempt"
            
    return df_uzio

def inject_into_uzio_template(df_uzio, template_path="templates/Uzio_Census_Template.xlsm"):
    """
    Injects a formatted Uzio DataFrame into the standard Uzio .xlsm template.
    Preserves all sheets, instructions, and headers.
    Dynamically finds the row containing 'Employee First Name*' and starts data on the next row.
    """
    import openpyxl
    import os
    import re
    
    if isinstance(template_path, str):
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template file not found at {template_path}")
        
    wb = openpyxl.load_workbook(template_path, keep_vba=True)
    ws = wb['Employee Details']
    
    # Dynamically find the header row
    header_row = 4 # Fallback
    headers_in_template = {}
    
    for r in range(1, 10): # Search first 10 rows
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val and re.sub(r'\s+', ' ', str(val)).strip() == 'Employee First Name*':
                header_row = r
                break
        if header_row == r:
            break
            
    # Map column headers
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val:
            # Normalize to handle templates with embedded newlines like 'Employment\nStatus*'
            norm_val = re.sub(r'\s+', ' ', str(val)).strip()
            headers_in_template[norm_val] = col_idx

    # Write data starting at the row after the headers
    start_row = header_row + 1

    for row_idx, row_data in df_uzio.iterrows():
        excel_row = start_row + row_idx
        for col_name in df_uzio.columns:
            c_name_strip = re.sub(r'\s+', ' ', str(col_name)).strip()
            if c_name_strip in headers_in_template:
                col_idx = headers_in_template[c_name_strip]
                val = row_data[col_name]
                if pd.notna(val) and val != "":
                    ws.cell(row=excel_row, column=col_idx, value=val)
                    
    return wb

def validate_uzio_data(df_uzio):
    """
    Validates required fields for Uzio Census.
    Returns a DataFrame containing Employee ID and the list of missing fields.
    Fields checked: Pay Type*, Employment Status*, Job Title, Work Location.
    """
    errors = []
    
    # Identify expected column names from UZIO_RAW_MAPPING vs what's in df_uzio
    # Or just use the exact Uzio headers if df_uzio has them
    emp_id_col = 'Employee ID*' if 'Employee ID*' in df_uzio.columns else 'Employee ID'
    
    for idx, row in df_uzio.iterrows():
        emp_id = row.get(emp_id_col, f"Row {idx+1}")
        if pd.isna(emp_id) or str(emp_id).strip() == "":
            emp_id = f"Row {idx+1}"
            
        missing_fields = []
        
        # Check Pay Type
        val_pt = row.get('Pay Type*')
        if pd.isna(val_pt) or str(val_pt).strip() == "":
            missing_fields.append("Pay Type")
            
        # Check Employment Status
        val_es = row.get('Employment Status*')
        if pd.isna(val_es) or str(val_es).strip() == "":
            missing_fields.append("Employment Status")
            
        # Check Job Title
        val_jt = row.get('Job Title')
        invalid_jt = False
        allowed_titles = [
            'dsp owner', 'operations manager', 'operations lead', 'fleet manager', 
            'safety manager', 'performance manager', 'trainer', 'human resources', 
            'recruiter', 'office personnel', 'payroll assistant', 'finance', 
            'dispatch', 'management', 'admin', 'survey', 'warehouse', 'walker', 
            'driver', 'helper', 'driver-lite', 'driver-step van', 
            'driver-unscheduled', 'lead driver', 'ddu dedicated', 'ddu shared', 
            'non-dsp related', 'driver -major appliance'
        ]
        
        if pd.isna(val_jt) or str(val_jt).strip() == "":
            missing_fields.append("Job Title")
        elif str(val_jt).strip().lower() not in allowed_titles:
            invalid_jt = True
            
        # Check Work Location
        val_wl = row.get('Work Location')
        if pd.isna(val_wl) or str(val_wl).strip() == "":
            missing_fields.append("Work Location")
            
        if missing_fields or invalid_jt:
            err_reasons = []
            if missing_fields:
                err_reasons.append("Mandatory fields are blank")
            if invalid_jt:
                err_reasons.append(f"Invalid Job Title: '{val_jt}'")
                
            errors.append({
                "Employee ID": emp_id,
                "Missing Fields": ", ".join(missing_fields),
                "Error": " | ".join(err_reasons)
            })
            
    return pd.DataFrame(errors)
    return pd.DataFrame(errors)

def read_uzio_template_df(file):
    """
    Reads the 'Employee Details' sheet from a Uzio template .xlsm.
    Identifies the header row (index 3) and returns the full DataFrame.
    """
    try:
        # Load the whole workbook to preserve everything, but read as DF for logic
        df_template = pd.read_excel(file, sheet_name='Employee Details', header=3, dtype=str)
        # Normalize column names for internal matching (strip space, handles newlines)
        df_template.columns = [str(c).replace("\n", " ").replace("\r", " ").strip() for c in df_template.columns]
        return df_template
    except Exception as e:
        print(f"Error reading Uzio template: {e}")
        return None

def selective_update_uzio(df_source, df_template, selected_uzio_cols, vendor_field_map, fix_options=None):
    """
    Updates specific columns in df_template using data from df_source.
    Only updates employees present in df_source.
    Returns the updated df_template and a summary of changes.
    """
    # 1. Reverse the mapping to find which 'standard field' relates to the selected 'Uzio Column'
    # selected_uzio_cols are the raw keys from UZIO_RAW_MAPPING
    
    # 2. Normalize Employee IDs for matching
    emp_id_col_source = vendor_field_map.get('Employee ID')
    emp_id_col_template = 'Employee ID*' if 'Employee ID*' in df_template.columns else 'Employee ID'
    
    if not emp_id_col_source or emp_id_col_source not in df_source.columns:
        return df_template, "Error: Source 'Employee ID' column not found."
    
    # Prepare lookup dict from source: {id: row_data}
    df_source_clean = df_source.copy()
    df_source_clean[emp_id_col_source] = norm_key_series(df_source_clean[emp_id_col_source])
    # --- FIX: Deduplicate to prevent "DataFrame index must be unique" error ---
    df_source_clean = df_source_clean.drop_duplicates(subset=[emp_id_col_source], keep='first')
    source_lookup = df_source_clean.set_index(emp_id_col_source).to_dict('index')
    
    # 3. Create a temp copy for formatting
    # We use a dummy generator logic to get formatted values for each standard field
    from utils.audit_utils import UZIO_RAW_MAPPING
    
    updated_count = 0
    df_updated = df_template.copy()
    
    # Track changes for display
    change_details = []

    for idx, row in df_updated.iterrows():
        eid = norm_key_series(pd.Series([row.get(emp_id_col_template, "")])).iloc[0]
        
        if eid in source_lookup:
            source_row_data = source_lookup[eid]
            # Create a mini-dataframe for this person to use existing formatters if needed
            # Or just pull logic from generate_uzio_template
            
            for uzio_col in selected_uzio_cols:
                std_name = UZIO_RAW_MAPPING.get(uzio_col)
                if not std_name: continue
                
                vendor_col = vendor_field_map.get(std_name)
                if not vendor_col or vendor_col not in df_source.columns: continue
                
                val = source_row_data.get(vendor_col)
                # Apply same formatting used in generate_uzio_template
                formatted_val = ""
                if pd.notna(val) and str(val).strip() != "":
                    # Reuse specific formatters
                    if std_name == 'Middle Initial':
                        formatted_val = str(val).strip()[0]
                    elif std_name in ['Hire Date', 'Original Hire Date', 'Termination Date', 'DOB']:
                        try:
                            dt = pd.to_datetime(str(val).strip(), errors='coerce')
                            formatted_val = dt.strftime('%d/%m/%Y') if not pd.isna(dt) else str(val).strip()
                        except:
                            formatted_val = str(val).strip()
                    elif std_name == 'License Expiration Date':
                        v_str = str(val).strip()
                        if fix_options and fix_options.get('fix_license', False):
                            if '00/00/0000' in v_str or v_str in ('0', '00', '0000'):
                                formatted_val = ""
                            else:
                                try:
                                    dt = pd.to_datetime(v_str, errors='coerce')
                                    formatted_val = dt.strftime('%m/%d/%Y') if not pd.isna(dt) else ""
                                except:
                                    formatted_val = ""
                        else:
                            formatted_val = v_str
                    elif std_name == 'SSN':
                        formatted_val = str(val).replace("-", "").strip()
                    elif std_name == 'Gender':
                        g_str = str(val).lower()
                        if g_str.startswith('m'): formatted_val = "Male"
                        elif g_str.startswith('f'): formatted_val = "Female"
                    elif std_name == 'Employment Status':
                        s = str(val).lower()
                        if fix_options and fix_options.get('fix_status', False):
                            if 'not hired' in s: formatted_val = 'EXCLUDE'
                            elif 'inactive' in s or 'term' in s: formatted_val = 'TERMINATED'
                            elif 'active' in s or 'leave' in s: formatted_val = 'ACTIVE'
                            else: formatted_val = str(val).strip().upper()
                        else:
                            formatted_val = str(val).strip().upper()
                    elif std_name in ['Zip', 'Mailing Zip']:
                        z_clean = re.sub(r'\D', '', str(val).strip())
                        formatted_val = z_clean.zfill(5)[:5] if z_clean else ""
                    elif std_name == 'Employment Type':
                        et_str = str(val).lower()
                        if fix_options and fix_options.get('fix_type', False):
                            if 'full' in et_str: formatted_val = 'Full Time'
                            elif 'part' in et_str: formatted_val = 'Part Time'
                            elif 'season' in et_str: formatted_val = 'Seasonal'
                            elif 'other' in et_str: formatted_val = 'Other'
                            elif 'intern' in et_str: formatted_val = 'Part Time'
                        else:
                            formatted_val = str(val).strip()
                    elif std_name == 'Termination Reason':
                        tr_str = str(val).strip().lower()
                        if "involuntary" in tr_str or "invluntary" in tr_str:
                            formatted_val = "Involuntary Termination of Employment"
                        elif "voluntary" in tr_str or "quit" in tr_str:
                            formatted_val = "Voluntary Termination of Employment"
                        elif "death" in tr_str:
                            formatted_val = "Death"
                        elif "retire" in tr_str:
                            formatted_val = "Retirement"
                        elif "disability" in tr_str:
                            formatted_val = "Permanent Disability"
                        elif "transfer" in tr_str:
                            formatted_val = "Transfer"
                        else:
                            formatted_val = "Other"
                    else:
                        formatted_val = str(val).strip()
                
                # Check for delete/update
                old_val = row.get(uzio_col, "")
                if str(formatted_val) != str(old_val):
                    df_updated.at[idx, uzio_col] = formatted_val
                    change_details.append({
                        'Employee ID': eid,
                        'Column': uzio_col,
                        'From': old_val,
                        'To': formatted_val
                    })
                    updated_count += 1

            # Special case: Email Fallback (only if Work Email was selected and fix_emails is True)
            if fix_options and fix_options.get('fix_emails', False):
                work_email_col = next((c for c in selected_uzio_cols if UZIO_RAW_MAPPING.get(c) == 'Work Email'), None)
                if work_email_col:
                    current_work_email = df_updated.at[idx, work_email_col]
                    if pd.isna(current_work_email) or str(current_work_email).strip() == "":
                        pers_email_col = next((c for c in df_source.columns if norm_colname(c).casefold() == 'personal email'), None)
                        if not pers_email_col:
                             # Try fuzzy match if direct fails
                             pers_email_col = next((c for c in df_source.columns if 'personal' in c.lower() and 'email' in c.lower()), None)
                        
                        if pers_email_col:
                            fallback_email = str(source_row_data.get(pers_email_col, "")).strip()
                            if fallback_email:
                                df_updated.at[idx, work_email_col] = fallback_email
                                summary_id = f"Email Fallback ({eid})"
                                if not any(d['Employee ID'] == eid and d['Column'] == work_email_col for d in change_details):
                                    change_details.append({
                                        'Employee ID': eid,
                                        'Column': work_email_col,
                                        'From': '(blank)',
                                        'To': fallback_email
                                    })
                                    updated_count += 1

    # --- Post-processing: License Rules on the updated template (Optional) ---
    if fix_options and fix_options.get('fix_license', False):
        lic_num_col = 'License Number*'
        lic_exp_col = 'License Expiration Date'
        if lic_exp_col in df_updated.columns:
            # Clear any 00/00/0000 or invalid placeholders
            bad_mask = df_updated[lic_exp_col].astype(str).str.strip().isin(['00/00/0000', '0', '00', '0000', 'nan', 'NaT', ''])
            df_updated.loc[bad_mask, lic_exp_col] = ""
            # Clear expiration date when license number is blank
            if lic_num_col in df_updated.columns:
                no_lic_mask = df_updated[lic_num_col].isna() | (df_updated[lic_num_col].astype(str).str.strip() == "") | (df_updated[lic_num_col].astype(str).str.strip() == 'nan')
                df_updated.loc[no_lic_mask, lic_exp_col] = ""

    summary = f"Updated {updated_count} employees. Total {len(change_details)} cell changes."
    return df_updated, summary, pd.DataFrame(change_details)
