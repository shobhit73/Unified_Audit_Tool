import streamlit as st
import pandas as pd
import io
import re
import openpyxl
from openpyxl.utils import get_column_letter

APP_TITLE = "Paycom vs Uzio – Time Off Tool"

def clean_id(x):
    """Normalize Employee ID (remove .0, strip, remove leading zeros)."""
    if pd.isna(x): return ""
    s = str(x).strip()
    if s.endswith(".0"): s = s[:-2]
    # Remove leading zeros to match typically
    s = s.lstrip("0")
    return s

def run_tool(file_paycom, file_uzio):
    # 1. Read Paycom Report (Likely HTML disguised as XLS)
    try:
        # Try read_html first as it's common for Paycom 'xls'
        dfs = pd.read_html(file_paycom, header=0)
        if not dfs:
            st.error("No tables found in Paycom file.")
            return None, None
        df_p = dfs[0] # Assume main table is first
    except ValueError:
        # Fallback if actual Excel or CSV
        file_paycom.seek(0)
        try:
             df_p = pd.read_excel(file_paycom)
        except:
             file_paycom.seek(0)
             df_p = pd.read_csv(file_paycom)
    except Exception as e:
        st.error(f"Error reading Paycom file: {e}")
        return None, None

    # Normalize Paycom Columns
    # Look for 'Employee Code' and 'Net Available'
    p_cols = {c.strip(): c for c in df_p.columns}
    
    col_id_p = next((c for c in p_cols if "Employee Code" in c or "Employee ID" in c or "EECode" in c), None)
    col_bal_p = next((c for c in p_cols if "Net Available" in c), None)

    if not col_id_p or not col_bal_p:
        st.error(f"Could not find required columns in Paycom file. Found: {list(df_p.columns)}")
        return None, None

    # Create Lookup Map: CleanID -> Net Available
    # Filter out rows where Net Available is NaN/Blank if we want to preserve blanks?
    # Requirement: "Keep them as blank operating balance... dont fill anything"
    # So we only map values that exist.
    
    balance_map = {}
    for idx, row in df_p.iterrows():
        eid = clean_id(row[col_id_p])
        val = row[col_bal_p]
        
        if eid and pd.notna(val) and str(val).strip() != "":
            balance_map[eid] = val

    # ---------------------------------------------------------
    # PART A: Generate Clean Import File (using openpyxl)
    # ---------------------------------------------------------
    file_uzio.seek(0) # Reset file pointer for openpyxl
    try:
        wb_import = openpyxl.load_workbook(file_uzio)
    except Exception as e:
        st.error(f"Error reading Uzio Template with openpyxl: {e}")
        return None, None

    # Sheet 2 is index 1
    if len(wb_import.sheetnames) < 2:
        st.error("Uzio Template must have at least 2 sheets (Instruction, Time Off Details).")
        return None, None
        
    ws = wb_import.worksheets[1] # "Time Off Details"
    
    # Header is Row 4. Data starts Row 5.
    header_row = 4
    
    # Identify Columns in Header Row
    # openpyxl uses 1-based indexing for rows/cols
    # Iterate through header row to find column indices
    idx_id_u = None
    idx_bal_u = None
    
    for cell in ws[header_row]:
        val = str(cell.value).strip() if cell.value else ""
        if "Employee ID" in val:
            idx_id_u = cell.column # 1-based index
        elif "Operating Balance" in val or "Opening Balance" in val:
            idx_bal_u = cell.column
            
    if not idx_id_u or not idx_bal_u:
        st.error(f"Could not find 'Employee ID' or 'Opening Balance' headers in Row 4 of Sheet 2.")
        return None, None

    # Iterate Data Rows
    for row_idx in range(header_row + 1, ws.max_row + 1):
        cell_id = ws.cell(row=row_idx, column=idx_id_u)
        cell_bal = ws.cell(row=row_idx, column=idx_bal_u)
        
        current_val = cell_bal.value
        
        # Rule: If Blank -> Keep Blank
        if current_val is None or str(current_val).strip() == "":
            continue # Skip
            
        # Policy Assigned -> Update
        eid = clean_id(cell_id.value)
        if eid in balance_map:
            cell_bal.value = balance_map[eid]
            
    # Save Import File
    out_import = io.BytesIO()
    wb_import.save(out_import)
    out_import.seek(0)

    # ---------------------------------------------------------
    # PART B: Generate Audit Report (using pandas)
    # ---------------------------------------------------------
    file_uzio.seek(0) # Reset file pointer again for pandas
    # Re-read for pandas processing
    try:
        df_u = pd.read_excel(file_uzio, sheet_name=1, header=3)
    except Exception as e:
        st.error(f"Error reading Uzio Template for audit: {e}")
        return out_import.getvalue(), None # Return at least import file
        
    u_cols = {c.strip(): c for c in df_u.columns}
    col_id_u = next((c for c in u_cols if "Employee ID" in c), None)
    col_bal_u = next((c for c in u_cols if "Operating Balance" in c), None)
    if not col_bal_u:
        col_bal_u = next((c for c in u_cols if "Opening Balance" in c), None)

    if not col_id_u or not col_bal_u:
        st.error(f"Could not find 'Employee ID' or 'Opening/Operating Balance' in Uzio Template for audit. Found: {list(df_u.columns)}")
        return out_import.getvalue(), None

    # Future Columns
    col_future_app = next((c for c in p_cols if "Future Approved" in c), None)
    col_future_pend = next((c for c in p_cols if "Future Pending" in c), None)

    # Trackers
    matched_paycom_ids = set()
    unassigned_policies_rows = [] 

    # function to apply map for audit tracking
    def update_balance_pd(row):
        current_val = row[col_bal_u]
        if pd.isna(current_val) or str(current_val).strip() == "":
            unassigned_policies_rows.append(row.to_dict())
            return current_val
        eid = clean_id(row[col_id_u])
        if eid in balance_map:
            matched_paycom_ids.add(eid)
            return balance_map[eid]
        return current_val

    df_u[col_bal_u] = df_u.apply(update_balance_pd, axis=1)

    # Additional Reports
    missing_in_uzio = []
    for idx, row in df_p.iterrows():
        eid = clean_id(row[col_id_p])
        if eid and eid not in matched_paycom_ids:
            val = row[col_bal_p]
            if pd.notna(val) and str(val).strip() != "":
                missing_in_uzio.append(row)
    
    df_missing = pd.DataFrame(missing_in_uzio)
    df_unassigned = pd.DataFrame(unassigned_policies_rows)

    future_rows = []
    if col_future_app and col_future_pend:
        for idx, row in df_p.iterrows():
            try:
                fa = float(row[col_future_app]) if pd.notna(row[col_future_app]) else 0
                fp = float(row[col_future_pend]) if pd.notna(row[col_future_pend]) else 0
                if fa > 0 or fp > 0:
                    future_rows.append(row)
            except:
                pass
    df_future = pd.DataFrame(future_rows)

    out_audit = io.BytesIO()
    with pd.ExcelWriter(out_audit, engine='xlsxwriter') as writer:
        df_u.to_excel(writer, sheet_name='Time Off Details (Updated)', index=False)
        if not df_missing.empty:
            df_missing.to_excel(writer, sheet_name='Missing in Uzio', index=False)
        else:
            pd.DataFrame({'Message': ['All Paycom employees matched']}).to_excel(writer, sheet_name='Missing in Uzio', index=False)
        if not df_unassigned.empty:
            df_unassigned.to_excel(writer, sheet_name='Unassigned Policies', index=False)
        else:
            pd.DataFrame({'Message': ['No unassigned policies found']}).to_excel(writer, sheet_name='Unassigned Policies', index=False)
        if not df_future.empty:
            df_future.to_excel(writer, sheet_name='Future Time Off', index=False)
        else:
            pd.DataFrame({'Message': ['No future time off found']}).to_excel(writer, sheet_name='Future Time Off', index=False)
        df_p.to_excel(writer, sheet_name='Paycom Raw Data', index=False)

    return out_import.getvalue(), out_audit.getvalue()

def render_ui():
    st.title(APP_TITLE)
    st.markdown("""
    **Instructions**:
    1. Upload **Paycom TimeOff Summary Report** (.xls / HTML).
    2. Upload **Uzio Time Off Import Template** (.xlsx).
    
    **Outputs**:
    - **Import File**: Clean file ready for Uzio import (preserves template).
    - **Audit Report**: Detailed analysis (Missing, Unassigned, Future).
    """)

    col1, col2 = st.columns(2)
    with col1:
        f_p = st.file_uploader("Paycom TimeOff Report", type=["xls", "html", "xlsx"], key="pt_p")
    with col2:
        f_u = st.file_uploader("Uzio Template", type=["xlsx"], key="pt_u")

    if st.button("Generate Reports", key="run_timeoff"):
        if not f_u or not f_p:
            st.error("Please upload both files.")
            return
            
        try:
            with st.spinner("Processing..."):
                res_import, res_audit = run_tool(f_p, f_u)
                
            if res_import and res_audit:
                st.success("Files Generated Successfully!")
                
                c1, c2 = st.columns(2)
                with c1:
                    st.download_button(
                        "Download Uzio Import File",
                        data=res_import,
                        file_name="Uzio_TimeOff_Import_Ready.xlsx"
                    )
                with c2:
                        st.download_button(
                        "Download Detailed Audit Report",
                        data=res_audit,
                        file_name="Uzio_TimeOff_Audit_Report.xlsx"
                    )
            elif res_import: # Only import file was generated
                st.warning("Audit report could not be generated, but import file is available.")
                st.download_button(
                    "Download Uzio Import File",
                    data=res_import,
                    file_name="Uzio_TimeOff_Import_Ready.xlsx"
                )
            else:
                st.error("No files could be generated due to an error.")

        except Exception as e:
            st.error(f"An error occurred: {e}")
            st.exception(e)

# Streamlit UI
if __name__ == "__main__":
    render_ui()
