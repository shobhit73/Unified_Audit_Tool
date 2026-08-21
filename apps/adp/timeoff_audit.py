import io
import re

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.utils.dataframe import dataframe_to_rows

from apps.adp.prior_payroll_sanity import _evaluate_cell

APP_TITLE = "ADP vs Uzio – Time Off Tool"

# Both UZIO workbooks (Time Off Import, Employee Census) put their headers on
# row 4 with a title/notes block above.
UZIO_HEADER_ROW = 4


def clean_id(x):
    """Normalize Employee ID (remove .0, strip, remove leading zeros)."""
    if pd.isna(x):
        return ""
    s = str(x).strip()
    if s.endswith(".0"):
        s = s[:-2]
    # Remove leading zeros to match typically
    s = s.lstrip("0")
    return s


def _norm_header(c):
    """UZIO headers carry stray spaces, embedded newlines and required-field
    asterisks (' Employee ID*', 'Employment\\nStatus*'). Flatten them so a
    lookup by plain name works."""
    return re.sub(r"\s+", " ", str(c)).strip().rstrip("*").strip().lower()


def _find(cols, *needles):
    """First column whose normalized header contains every needle."""
    for c in cols:
        n = _norm_header(c)
        if all(x in n for x in needles):
            return c
    return None


# ─────────────────────────────────────────────────────────────────────────────
# ADP side
# ─────────────────────────────────────────────────────────────────────────────

def read_adp_balances(file_adp):
    """One row per employee: total time-off balance across their policy rows.

    ADP writes money cells as `=ROUND(x, 2.0)` formulas WITHOUT caching the
    computed value, so `pd.read_excel` (which reads the cache) sees the whole
    BALANCE AMOUNT column as null. Summing an all-null group then yields 0.0
    rather than NaN, which silently filled every Opening Balance with zero. Read
    through openpyxl and evaluate the formulas instead — the same treatment
    `prior_payroll_sanity` already gives ADP money columns.

    ADP's "Totals For <name> - <policy> -- Balance Amount:" subtotal rows carry
    no ASSOCIATE ID, so dropping blank IDs removes them and the per-transaction
    rows sum to exactly ADP's own subtotal.

    Returns (DataFrame[id, name, balance], error).
    """
    try:
        wb = openpyxl.load_workbook(file_adp, data_only=False)
    except Exception as e:
        return None, f"Error reading ADP file: {e}"

    ws = None
    for sheet in wb.worksheets:
        head = [sheet.cell(row=1, column=c).value for c in range(1, sheet.max_column + 1)]
        if _find(head, "associate id") and _find(head, "balance amount"):
            ws = sheet
            break
    if ws is None:
        return None, ("Could not find a sheet with both `ASSOCIATE ID` and "
                      "`BALANCE AMOUNT` columns in the ADP file.")

    head = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(head)}
    c_id = idx[_find(head, "associate id")]
    c_bal = idx[_find(head, "balance amount")]
    name_hdr = next((h for h in head
                     if h and "name" in _norm_header(h) and "policy" not in _norm_header(h)), None)
    c_name = idx[name_hdr] if name_hdr else None

    rows = []
    for r in range(2, ws.max_row + 1):
        eid = clean_id(ws.cell(row=r, column=c_id).value)
        if not eid:
            continue          # subtotal / blank rows
        rows.append({
            "id": eid,
            "name": ws.cell(row=r, column=c_name).value if c_name else "N/A",
            "balance": _evaluate_cell(ws.cell(row=r, column=c_bal).value),
        })
    if not rows:
        return None, "No valid Employee IDs found in ADP file."

    df = pd.DataFrame(rows)
    df["balance"] = pd.to_numeric(df["balance"], errors="coerce")
    out = (df.groupby("id")
             .agg(balance=("balance", _sum_money), name=("name", "first"))
             .reset_index())
    return out, None


def _sum_money(series):
    """Total a set of 2-decimal money values back into a 2-decimal money value.

    `min_count=1` keeps an employee with no readable amounts at NaN rather than
    reporting a real 0.00 balance.

    The rounding matters: the transaction amounts are exact to the cent as
    decimals, but not in binary floating point, so adding them leaves dust.
    3.62 + 24.42 - 28.04 is 0.00 on paper and 3.55e-15 in a float. On this
    client 71 of 132 employees carried such dust; it stayed invisible wherever
    the total was large (18.709999999999994 renders as 18.71) and surfaced only
    where the total was exactly zero, which Excel then showed as `3.55271E-15`.
    """
    total = series.sum(min_count=1)
    return total if pd.isna(total) else round(total, 2)


# ─────────────────────────────────────────────────────────────────────────────
# UZIO census
# ─────────────────────────────────────────────────────────────────────────────

def read_census(file_census):
    """Employee ID → employment status + termination date from a UZIO census.

    Returns (DataFrame[id, status, termination date, name], error).
    """
    try:
        book = pd.read_excel(file_census, sheet_name=None,
                             header=UZIO_HEADER_ROW - 1, dtype=str)
    except Exception as e:
        return None, f"Error reading census file: {e}"

    for df in book.values():
        c_id = _find(df.columns, "employee id")
        c_st = _find(df.columns, "employment", "status")
        if c_id and c_st:
            c_td = _find(df.columns, "termination date")
            c_fn = _find(df.columns, "employee first name")
            c_ln = _find(df.columns, "employee last name")
            out = pd.DataFrame({
                "id": df[c_id].apply(clean_id),
                "status": df[c_st].fillna("").astype(str).str.strip(),
                "termination date": (df[c_td].fillna("").astype(str).str.strip()
                                     if c_td else ""),
                "census name": ((df[c_fn].fillna("") + " " + df[c_ln].fillna("")).str.strip()
                                if c_fn and c_ln else ""),
            })
            return out[out["id"] != ""].reset_index(drop=True), None

    return None, ("Could not find `Employee ID` and `Employment Status` columns "
                  f"on row {UZIO_HEADER_ROW} of any sheet in the census file.")


# ─────────────────────────────────────────────────────────────────────────────
# File 1 — the filled import template
# ─────────────────────────────────────────────────────────────────────────────

def fill_import_template(file_uzio, balance_map):
    """Write ADP balances into a copy of the UZIO Time Off Import template.

    The workbook is returned untouched apart from the Opening Balance column —
    same sheets, same formatting — so it can be uploaded to UZIO as-is.

    A BLANK Opening Balance means no policy is assigned to that employee, so it
    is left blank rather than filled.

    Returns (workbook, filled_count, error).
    """
    file_uzio.seek(0)
    try:
        wb = openpyxl.load_workbook(file_uzio)
    except Exception as e:
        return None, 0, f"Error reading Uzio Template: {e}"

    ws = next((s for s in wb.worksheets
               if _find([s.cell(row=UZIO_HEADER_ROW, column=c).value
                         for c in range(1, s.max_column + 1)], "employee id")), None)
    if ws is None:
        return None, 0, (f"Could not find an `Employee ID` header on row "
                         f"{UZIO_HEADER_ROW} of any sheet in the Uzio template.")

    head = [ws.cell(row=UZIO_HEADER_ROW, column=c).value for c in range(1, ws.max_column + 1)]
    pos = {h: i + 1 for i, h in enumerate(head)}
    h_id = _find(head, "employee id")
    h_bal = _find(head, "opening balance") or _find(head, "operating balance")
    if not h_id or not h_bal:
        return None, 0, ("Could not find `Employee ID` or `Opening Balance` headers "
                         f"in row {UZIO_HEADER_ROW} of the Uzio template.")

    filled = 0
    for r in range(UZIO_HEADER_ROW + 1, ws.max_row + 1):
        cell_bal = ws.cell(row=r, column=pos[h_bal])
        if cell_bal.value is None or str(cell_bal.value).strip() == "":
            continue                                    # unassigned policy
        eid = clean_id(ws.cell(row=r, column=pos[h_id]).value)
        if eid in balance_map and pd.notna(balance_map[eid]):
            cell_bal.value = balance_map[eid]
            filled += 1
    return wb, filled, None


# ─────────────────────────────────────────────────────────────────────────────
# File 2 — the audit workbook
# ─────────────────────────────────────────────────────────────────────────────

def build_audit_sheets(file_uzio, adp_df, census_df):
    """Every audit view, as {sheet name: DataFrame}.

    An employee whose template row exists but has a BLANK balance is an
    unassigned policy, NOT missing from UZIO — both used to be reported, so the
    same person appeared twice. Template rows are matched regardless of whether
    the balance is filled, and only genuinely absent IDs count as missing.
    """
    file_uzio.seek(0)
    df_u = pd.read_excel(file_uzio, sheet_name="Time Off Details",
                         header=UZIO_HEADER_ROW - 1)
    c_id = _find(df_u.columns, "employee id")
    c_bal = _find(df_u.columns, "opening balance") or _find(df_u.columns, "operating balance")
    c_name = _find(df_u.columns, "employee first name")

    balance_map = dict(zip(adp_df["id"], adp_df["balance"]))
    name_map = dict(zip(adp_df["id"], adp_df["name"]))

    template_ids, unassigned_rows, exceptions = set(), [], []
    for _, row in df_u.iterrows():
        eid = clean_id(row[c_id])
        if eid:
            template_ids.add(eid)                       # present, filled or not
        val = row[c_bal]
        if pd.isna(val) or str(val).strip() == "":
            unassigned_rows.append(row.to_dict())
            exceptions.append({
                "Employee ID": str(row[c_id]) if pd.notna(row[c_id]) else "",
                "Employee Name": str(row[c_name]) if c_name and pd.notna(row[c_name]) else "N/A",
                "Issue Category": "Unassigned Policy (Blank Balance)",
                "ADP Balance": "",
            })

    missing = []
    for _, row in adp_df.iterrows():
        eid, val = row["id"], row["balance"]
        if eid in template_ids or pd.isna(val):
            continue
        missing.append({"Employee ID": eid, "Employee Name": row["name"],
                        "Total Balance": val})
        exceptions.append({"Employee ID": eid, "Employee Name": row["name"],
                           "Issue Category": "Missing in Uzio Template",
                           "ADP Balance": val})

    sheets = {}

    if census_df is not None:
        status = dict(zip(census_df["id"], census_df["status"]))
        termdt = dict(zip(census_df["id"], census_df["termination date"]))
        rows = []
        for eid, bal in balance_map.items():
            st_val = str(status.get(eid, "")).strip()
            rows.append({
                "Employee ID": eid,
                "Employee Name": name_map.get(eid, "N/A"),
                "ADP Balance": bal,
                "UZIO Employment Status": st_val or "(not in census)",
                "Termination Date": termdt.get(eid, ""),
                "In Import Template": "Yes" if eid in template_ids else "No",
            })
        df_status = pd.DataFrame(rows)
        # Terminated first — those are the rows to act on.
        df_status["_rank"] = df_status["UZIO Employment Status"].str.lower().map(
            lambda s: 0 if s.startswith("terminated") else (2 if s == "active" else 1))
        df_status = (df_status.sort_values(["_rank", "ADP Balance"], ascending=[True, False])
                     .drop(columns="_rank").reset_index(drop=True))
        sheets["Balance vs UZIO Status"] = df_status

        for _, r in df_status[df_status["UZIO Employment Status"]
                              .str.lower().str.startswith("terminated")].iterrows():
            exceptions.append({
                "Employee ID": r["Employee ID"], "Employee Name": r["Employee Name"],
                "Issue Category": "Terminated in UZIO but ADP sent a balance",
                "ADP Balance": r["ADP Balance"],
            })

    sheets["Missing in Uzio"] = (pd.DataFrame(missing) if missing
                                 else pd.DataFrame({"Message": ["All ADP employees matched"]}))
    sheets["Unassigned Policies"] = (pd.DataFrame(unassigned_rows) if unassigned_rows
                                     else pd.DataFrame({"Message": ["No unassigned policies found"]}))
    sheets["ADP Grouped Data"] = adp_df.rename(
        columns={"id": "Employee ID", "name": "Employee Name", "balance": "Total Balance"})
    sheets["Exception Summary"] = (pd.DataFrame(exceptions) if exceptions
                                   else pd.DataFrame({"Message": ["No exceptions found"]}))
    return sheets


def audit_workbook_bytes(sheets):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, df in sheets.items():
        ws = wb.create_sheet(title=name[:31])
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def run_tool(file_adp, file_uzio, file_census=None):
    """Returns (filled_template_bytes, audit_bytes, stats) or (None, None, None)."""
    adp_df, err = read_adp_balances(file_adp)
    if err:
        st.error(err)
        return None, None, None

    census_df = None
    if file_census is not None:
        census_df, err = read_census(file_census)
        if err:
            st.error(err)
            return None, None, None

    balance_map = dict(zip(adp_df["id"], adp_df["balance"]))
    wb_filled, filled, err = fill_import_template(file_uzio, balance_map)
    if err:
        st.error(err)
        return None, None, None

    sheets = build_audit_sheets(file_uzio, adp_df, census_df)

    buf = io.BytesIO()
    wb_filled.save(buf)

    stats = {
        "employees": len(adp_df),
        "filled": filled,
        "missing": len(sheets["Missing in Uzio"]) if "Employee ID" in sheets["Missing in Uzio"] else 0,
        "unassigned": (len(sheets["Unassigned Policies"])
                       if "Message" not in sheets["Unassigned Policies"] else 0),
        "terminated": (int(sheets["Balance vs UZIO Status"]["UZIO Employment Status"]
                           .str.lower().str.startswith("terminated").sum())
                       if "Balance vs UZIO Status" in sheets else 0),
    }
    return buf.getvalue(), audit_workbook_bytes(sheets), stats


def render_ui():
    st.title(APP_TITLE)
    client_name = st.text_input("Client Name", value="Client", key="adp_timeoff_client")

    st.markdown("""
    **Upload**
    1. **ADP Time Off Balance Summary** (.xlsx)
    2. **Uzio Time Off Import Template** (.xlsx)
    3. **UZIO Employee Census** (.xlsx / .xlsm) — optional

    **You get two separate files**
    - `<Client>_Time off Import_filled.xlsx` — your template, unchanged except the
      filled Opening Balance column. Upload it to UZIO as-is.
    - `<Client>_Uzio_ADP_TimeOff_Audit_Report_<timestamp>.xlsx` — the audit only.

    Add the census and the audit gains a **Balance vs UZIO Status** sheet showing
    every employee ADP sent a balance for, with their UZIO employment status and
    termination date — terminated employees first.
    """)

    col1, col2, col3 = st.columns(3)
    with col1:
        f_a = st.file_uploader("ADP Balance Summary", type=["xlsx"], key="at_a")
    with col2:
        f_u = st.file_uploader("Uzio Template", type=["xlsx"], key="at_u")
    with col3:
        f_c = st.file_uploader("UZIO Census (optional)", type=["xlsx", "xlsm"], key="at_c")

    # st.download_button triggers a rerun of its own, so results computed inside
    # the Generate block would vanish the moment the first file is downloaded —
    # taking the second download button with them. Keep them in session_state and
    # render the buttons OUTSIDE that block.
    SKEY = "adp_timeoff_result"

    def _signature(*files):
        return tuple((f.name, getattr(f, "size", None)) if f is not None else None
                     for f in files)

    sig = _signature(f_a, f_u, f_c)
    cached = st.session_state.get(SKEY)
    if cached and cached.get("signature") != sig:
        # Different uploads than the ones that produced these files — drop them
        # rather than let someone download the previous client's data.
        del st.session_state[SKEY]
        cached = None

    if st.button("Generate Files", key="run_timeoff_adp"):
        if not f_u or not f_a:
            st.error("Please upload both the ADP Balance Summary and the Uzio Template.")
            return
        try:
            with st.spinner("Processing..."):
                filled_bytes, audit_bytes, stats = run_tool(f_a, f_u, f_c)
            if not filled_bytes:
                return
            st.session_state[SKEY] = {
                "signature": sig,
                "filled": filled_bytes,
                "audit": audit_bytes,
                "stats": stats,
                "had_census": f_c is not None,
                # Stamped once, so the filename does not change on every rerun.
                "ts": pd.Timestamp.now().strftime("%d_%m_%Y_%H%M"),
            }
            cached = st.session_state[SKEY]
        except Exception as e:
            st.error(f"An error occurred: {e}")
            st.exception(e)
            return

    if not cached:
        return

    stats = cached["stats"]
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Employees in ADP", stats["employees"])
    c2.metric("Balances written", stats["filled"])
    c3.metric("Missing in UZIO", stats["missing"])
    c4.metric("Terminated in UZIO", stats["terminated"])
    if not cached["had_census"]:
        st.caption("Upload the UZIO census to see which of these employees are "
                   "already terminated.")

    st.success("Both files are ready — download them one at a time.")
    d1, d2 = st.columns(2)
    with d1:
        st.download_button(
            "⬇️ Time Off Import (filled)",
            data=cached["filled"],
            file_name=f"{client_name}_Time off Import_filled.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            key="to_dl_filled",
        )
    with d2:
        st.download_button(
            "⬇️ Audit Report",
            data=cached["audit"],
            file_name=(f"{client_name}_Uzio_ADP_TimeOff_Audit_Report_"
                       f"{cached['ts']}.xlsx"),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="to_dl_audit",
        )
