"""Verify both Streamlit and audit_fast_api versions write a 3-tab xlsx."""
import sys, io
sys.path.insert(0, r"c:\Users\shobhit.sharma\Downloads\Deduction Tool\audit_fast_api")
sys.path.insert(0, r"c:\Users\shobhit.sharma\Downloads\Deduction Tool")

import openpyxl

# audit_fast_api version
from core.adp.prior_payroll_setup_helper import (
    run_adp_prior_payroll_setup_helper, build_simplified_xlsx_bytes,
)
with open(r"C:\Users\shobhit.sharma\Downloads\State Tax Code.csv", "rb") as f:
    master = f.read()
with open(r"C:\Users\shobhit.sharma\Downloads\Carvan Prior Payroll Setup\Payroll_History_Q1_Consolidated.csv", "rb") as f:
    carvan = f.read()

results, _ = run_adp_prior_payroll_setup_helper(carvan, "Carvan_Q1.csv", master)
xlsx_bytes = build_simplified_xlsx_bytes(results)
out_path = r"C:\Users\shobhit.sharma\Desktop\Audit Files\TEST_simple_setup.xlsx"
with open(out_path, "wb") as f:
    f.write(xlsx_bytes)

wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
print(f"Tabs in audit_fast_api xlsx: {wb.sheetnames}")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n  --- {name} ({ws.max_row} rows x {ws.max_column} cols) ---")
    for row in list(ws.iter_rows(values_only=True))[:15]:
        print(f"    {row}")

# Streamlit version
print("\n\n========== STREAMLIT VERSION ==========")

class Buf(io.BytesIO):
    def __init__(self, c, name):
        super().__init__(c); self.name = name

# Reload via the Streamlit module path. Need streamlit available.
try:
    from apps.adp.prior_payroll_setup_helper import run_setup_helper, _results_to_xlsx_bytes
    adp = Buf(carvan, "Carvan_Q1.csv")
    mfile = Buf(master, "State Tax Code.csv")
    res2, _ = run_setup_helper(adp, mfile)
    xlsx2 = _results_to_xlsx_bytes(res2)
    wb2 = openpyxl.load_workbook(io.BytesIO(xlsx2))
    print(f"Tabs in Streamlit xlsx: {wb2.sheetnames}")
    for name in wb2.sheetnames:
        ws = wb2[name]
        print(f"\n  --- {name} ({ws.max_row} rows x {ws.max_column} cols) ---")
        for row in list(ws.iter_rows(values_only=True))[:15]:
            print(f"    {row}")
except Exception as e:
    print(f"Streamlit version test failed: {e}")
